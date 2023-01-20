# -*- coding: utf-8 -*-
"""
Created on Wed Sep 22 16:43:24 2021

@author: PathakS
"""

import re
import os
import ast
import sys
import math
import torch
import datetime
import argparse
import itertools
import random

import glob
import numpy as np
import pandas as pd
import openpyxl as op
import torch.nn as nn
from PIL import Image
from sklearn import metrics
#from skimage import io
import torch.optim as optim
from openpyxl import Workbook
#from torchsummary import summary
from configparser import ConfigParser
from torch import linalg as LA

from torchvision import models
#from torchviz import make_dot
import matplotlib.pyplot as plt
import torch.nn.functional as F
from torchvision import transforms
from pytorchtools import EarlyStopping
from sklearn.metrics import confusion_matrix
from torchvision.models.resnet import BasicBlock
from torch.utils.tensorboard import SummaryWriter
from sklearn.model_selection import StratifiedKFold
from sklearn.model_selection import train_test_split
from sklearn.model_selection import GroupShuffleSplit
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from torchvision.ops.focal_loss import sigmoid_focal_loss

import test
import utils
import mymodels
from model_checkpoint import ModelCheckpoint
#import tensorboard_log

#pd.set_option('display.max_columns', None)  # or 1000
#pd.set_option('display.max_rows', None)  # or 1000
pd.set_option('display.max_colwidth', None)  # or 199

def results_store_excel(train_res, val_res, test_res, per_model_metrics, correct_train,total_images_train,train_loss,correct_test,total_images_test,test_loss,epoch,conf_mat_train,conf_mat_test, lr, auc_val):
    lines = [epoch+1, lr]
    if train_res:
        avg_train_loss=train_loss/total_images_train
        accuracy_train=correct_train / total_images_train
        speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
        prec_train=conf_mat_train[1,1]/sum(conf_mat_train[:,1])
        f1_train=2*recall_train*prec_train/(recall_train+prec_train)
        prec_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[:,0])
        recall_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        f1_train_neg=2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
        f1macro_train=(f1_train+f1_train_neg)/2
        lines.extend([avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train])

    if val_res:
        speci_test=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        avg_test_loss=test_loss/total_images_test
        recall_test=conf_mat_test[1,1]/sum(conf_mat_test[1,:])
        prec_test=conf_mat_test[1,1]/sum(conf_mat_test[:,1])
        f1_test=2*recall_test*prec_test/(recall_test+prec_test)
        accuracy_test=correct_test / total_images_test
        recall_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        prec_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[:,0])
        f1_test_neg=2*recall_test_neg*prec_test_neg/(recall_test_neg+prec_test_neg)
        f1macro_test=(f1_test+f1_test_neg)/2
        lines.extend([avg_test_loss, accuracy_test, f1macro_test, recall_test, speci_test, auc_val])

    if test_res:
        lines.extend(per_model_metrics)
    
    #lines=[epoch+1, lr, avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train, avg_test_loss, accuracy_test, f1macro_test, recall_test, speci_test]
    out=open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    sheet1.append(lines)
    out.close()
    #tensorboard_log.loss_plot(writer, file_name, avg_train_loss, avg_test_loss, epoch)
    #tensorboard_log.acc_plot(writer, file_name, accuracy_train, accuracy_test, epoch)

def results_plot(df, file_name):
    plt.plot(df['Epoch'],df['F1macro Train'],'-r',label='Train F1macro')
    plt.plot(df['Epoch'],df['F1macro Val'],'-b',label='Val F1macro')
    plt.plot(df['Epoch'],df['Avg Loss Train'],'-g',label='Train Loss')
    plt.plot(df['Epoch'],df['Avg Loss Val'],'-y',label='Val Loss')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,df.iloc[-1]['Epoch']))
    plt.xlabel('Epochs')
    plt.ylabel('Loss/F1macro')
    plt.title('Learning curve')
    plt.savefig(path_to_learning_curve, bbox_inches='tight')

def conf_mat_create(predicted,true,correct,total_images,conf_mat):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images,conf_mat,conf_mat_batch

def save_model(model,optimizer,epoch, loss):
    state = {
        'epoch': epoch+1,
        'state_dict': model.state_dict(),
        'optim_dict': optimizer.state_dict(),
        'loss': loss
    }
    torch.save(state,path_to_model)

def load_model(model,optimizer,path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    optimizer.load_state_dict(checkpoint['optim_dict'])
    epoch = checkpoint['epoch']
    return model,optimizer,epoch

def load_model_for_testing(model,path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    print("checkpoint epoch and loss:", checkpoint['epoch'], checkpoint['loss'])
    return model 

'''def adaptive_learning_rate(optimizer, epoch, init_lr=0.001):
    """Sets the learning rate to the initial LR decayed by 0.2 every 10 epochs"""
    lr = init_lr * (0.2 ** (epoch // 10)) #0.001, 0.0002, 0.00004, 0.000008 
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr
    
    return optimizer'''

def lrdecay_scheduler(optimizer):#, epoch):
    """Sets the learning rate to the initial LR decayed by 0.2 every 10 epochs"""
    #for param_group in optimizer.param_groups:
    #    if not epoch%8:
    #        if epoch!=0:
    #            lr = param_group['lr'] * 0.98
    #            param_group['lr'] = lr
    scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=8, gamma=0.98)
    return scheduler

def multisteplr_routine1(optimizer): #multisteplr1
    scheduler = optim.lr_scheduler.MultiStepLR(optimizer, milestones=[epoch_step], gamma=0.1) #6 means 7. epochs starts with 0. So, from 7th epoch the model will move to another learning rate
    return scheduler

def optimizer_fn(data_type):
    if data_type=='variable':
        mlo_group=[]
        cc_group=[]
        both_attention_group=[]
        rest_group=[]
        param_list=[]
        
        for name,param in model.named_parameters():
            if param.requires_grad:
                if '.mlo' in name:
                    mlo_group.append(param)
                elif '.cc' in name:
                    cc_group.append(param) 
                elif '_both_b.attention' in name or '_both_m.attention' in name:
                    both_attention_group.append(param)
                else:
                    rest_group.append(param)
        for item in [mlo_group,cc_group,both_attention_group,rest_group]:
            if item:
                param_list.append({"params":item,"lr":0.001, "momentum":0.9, "weight_decay":wtdecay})
        if optimizer_type=='Adam':
            optimizer = optim.Adam(param_list)
        elif optimizer_type=='SGD':
            optimizer = optim.SGD(param_list)
    
    elif data_type=='fixed':
        classifier=[]
        rest_group=[]
        if optimizer_type=='Adam':
            optimizer = optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=lrval, weight_decay=wtdecay)
            '''for name,param in model.named_parameters():
                if param.requires_grad:
                    if '.fc' in name:
                        classifier.append(param)
                    else:
                        rest_group.append(param)
            optimizer = optim.Adam([{'params':classifier, 'lr':0.0001, "weight_decay":wtdecay },{'params':rest_group, 'lr':lrval, "weight_decay":wtdecay }])
            '''
        elif optimizer_type=='SGD':
            optimizer = optim.SGD(filter(lambda p: p.requires_grad, model.parameters()), lr=lrval, momentum=0.9, weight_decay=wtdecay)
    return optimizer

class WeightedFocalLoss(nn.Module):
    "Non weighted version of Focal Loss"
    def __init__(self, alpha=.25, gamma=2):
        super(WeightedFocalLoss, self).__init__()
        self.alpha = torch.tensor([alpha, 1-alpha]).cuda()
        self.gamma = gamma

    def forward(self, inputs, targets):
        BCE_loss = F.binary_cross_entropy_with_logits(inputs, targets, reduction='none')
        targets = targets.type(torch.long)
        at = self.alpha.gather(0, targets.data.view(-1))
        pt = torch.exp(-BCE_loss)
        F_loss = at*(1-pt)**self.gamma * BCE_loss
        return F_loss.mean()

def loss_fn(activation,class_imb,class_weights):
    if activation=='softmax':
        if class_imb=='wtcostfunc':
            criterion=nn.CrossEntropyLoss(class_weights)
        elif class_imb=='poswt':
            print("poswt weight:",torch.tensor([1,class_weights[0]]))
            if extra=='labelsmoothing':
                #criterion=nn.CrossEntropyLoss(weight=torch.tensor([1,class_weights[0]]).type(torch.float16).to(device), label_smoothing=0.1)
                criterion=nn.CrossEntropyLoss(weight=torch.tensor([1,class_weights[0]]).float().to(device), label_smoothing=0.1)
            else:
                #criterion=nn.CrossEntropyLoss(torch.tensor([1,class_weights[0]]).type(torch.float16).to(device))
                criterion=nn.CrossEntropyLoss(torch.tensor([1,class_weights[0]]).float().to(device))
        else:
            if extra=='labelsmoothing':
                criterion=nn.CrossEntropyLoss(label_smoothing=0.1)
            else:
                criterion=nn.CrossEntropyLoss()
    elif activation=='sigmoid':
        if class_imb=='poswt':
            criterion = nn.BCEWithLogitsLoss(pos_weight=class_weights[0])
        else:
            criterion = nn.BCEWithLogitsLoss()
    return criterion

def gmic_loss_fn(bcelogitsloss, bceloss, y_local, y_global, y_fusion, saliency_map, y_true, sm_reg_param):
    local_network_loss = bcelogitsloss(y_local, y_true)
    global_network_loss = bceloss(y_global, y_true)
    fusion_network_loss = bcelogitsloss(y_fusion, y_true)
    saliency_map_regularizer = torch.mean(LA.norm(saliency_map.view(saliency_map.shape[0],saliency_map.shape[1],-1), ord=1, dim=2))
    total_loss = local_network_loss + global_network_loss + fusion_network_loss + sm_reg_param*saliency_map_regularizer
    return total_loss

def train(model,data_iterator_train,data_iterator_test,batches_train,batches_test,epochs):
    '''Training'''
    optimizer = optimizer_fn(data_type)
    #scheduler = optim.lr_scheduler.MultiStepLR(optimizer, milestones=[4, 7], gamma=0.1)
    if patience_epoch:
        early_stopping = EarlyStopping(path_to_model=path_to_model,patience=patience_epoch,verbose=True)
    elif use_validation:
        modelcheckpoint = ModelCheckpoint(path_to_model=path_to_model,verbose=True)
    if os.path.isfile(path_to_model):
        #if trainingmethod=='multisteplr1':
        model,_,start_epoch=load_model(model,optimizer,path_to_model)
        #else:
        #    model,optimizer,start_epoch=load_model(model,optimizer,path_to_model)
        #    optimizer=adaptive_learning_rate(optimizer,start_epoch)
        print("start epoch:",start_epoch)
        print("lr:",optimizer.param_groups[0]['lr'])
    else:
        start_epoch=0
    
    if featureextractormodel=='gmic_resnet18_pretrained':
        if class_imbalance=='poswt':
            bcelogitloss = nn.BCEWithLogitsLoss(pos_weight=class_weights_train[0])
            bceloss = nn.BCELoss()
        else:
            bcelogitloss = nn.BCEWithLogitsLoss()
            bceloss = nn.BCELoss()
    else:
        if class_imbalance!='focalloss':
            lossfn = loss_fn(activation,class_imbalance,class_weights_train)
    
    if trainingmethod=='multisteplr1':
        scheduler = multisteplr_routine1(optimizer)
    elif trainingmethod=='lrdecayshu':
        scheduler = lrdecay_scheduler(optimizer)
    for epoch in range(start_epoch,epochs):
        model.train()
        #if trainingmethod==False:
        #    optimizer=adaptive_learning_rate(optimizer,epoch)
        
        loss_train=0.0
        correct_train=0
        conf_mat_train=np.zeros((2,2))
        total_images_train=0
        batch_no=0
        if trainingmethod=='multisteplr1':
            model=utils.layer_selection_for_training(model,epoch,trainingmethod,epoch_step)
        
        #print("lr last layer:",optimizer.param_groups[0]['lr'])
        
        for train_idx, train_batch, train_labels in data_iterator_train:
            #for name, param in model.named_parameters():
            #    if param.requires_grad:
            #        print(name)
            #input('halt')
            train_batch = train_batch.to(device)
            train_labels = train_labels.to(device)
            train_labels = train_labels.view(-1)#.float()
            print("train batch:",train_batch.shape)
            
            #if data_type=='variable':
            #    model, optimizer = utils.freeze_pipelines(model, optimizer, views_names, attention, feature_extractor)
            #with torch.cuda.amp.autocast():
            if featureextractormodel=='gmic_resnet18_pretrained':
                output_batch_local, output_batch_global, output_batch_fusion, saliency_map = model(train_batch) # compute model output, loss and total train loss over one epoch
                if class_imbalance=='poswt':
                    weight_batch = torch.tensor([1,class_weights_train[0]]).float().to(device)[train_labels]
                    bceloss.weight = weight_batch
                output_batch_local = output_batch_local.view(-1)
                output_batch_global = output_batch_global.view(-1)
                output_batch_fusion = output_batch_fusion.view(-1)
                train_labels = train_labels.float()
                pred = torch.ge(torch.sigmoid(output_batch_fusion), torch.tensor(0.5)).float()
                loss = gmic_loss_fn(bcelogitloss, bceloss, output_batch_local, output_batch_global, output_batch_fusion, saliency_map, train_labels, sm_reg_param)
            
            else:
                output_batch = model(train_batch)
                if activation=='sigmoid':
                    if class_imbalance=='wtcostfunc':
                        weight_batch = class_weights_train[train_labels]
                        lossfn.weight = weight_batch
                    output_batch = output_batch.squeeze(1)
                    output_batch = output_batch.view(-1)                                                                          
                    train_labels = train_labels.float()
                    pred = torch.ge(torch.sigmoid(output_batch), torch.tensor(0.5)).float()
                    #outputs = model(inputs)
                    #preds = torch.ge(torch.sigmoid(outputs), torch.tensor(0.5)).float()
                    #preds = preds.view(-1)
                    #outputs = outputs.squeeze(1)
                    #labels = labels.float()
                    #loss = criterion(outputs, labels)

                    if class_imbalance=='focalloss':
                        loss = sigmoid_focal_loss(output_batch, train_labels, alpha=-1, reduction='mean')
                    else:
                        loss = lossfn.forward(output_batch, train_labels)
                elif activation=='softmax':
                    pred = output_batch.argmax(dim=1, keepdim=True)
                    loss = lossfn(output_batch, train_labels)
            
            loss_train+=(train_labels.size()[0]*loss.item())
         
            
            '''weights_before_backprop = []
            parameter_name=[]

            for name, param in model.named_parameters(): # loop the weights in the model before updating and store them
                parameter_name.append(name)
                weights_before_backprop.append(param.clone())
            '''

            optimizer.zero_grad()  # clear previous gradients, compute gradients of all variables wrt loss
            loss.backward()
            optimizer.step() # performs updates using calculated gradients
            # Scales the loss, and calls backward()
            # to create scaled gradients
            #scaler.scale(loss).backward()
            # Unscales gradients and calls
            # or skips optimizer.step()
            #scaler.step(optimizer)
            # Updates the scale for next iteration
            #scaler.update()
            
            #loss.backward()
            #optimizer.step() # performs updates using calculated gradients
            batch_no=batch_no+1
            #for name, param in model.named_parameters():
            #    if param.requires_grad:
            #        tensorboard_log.grad_norm_plot(writer, name, param, epoch*batches_train+batch_no)
        
            utils.adam_state_access(optimizer)
            input('wait')

            '''weights_after_backprop = [] # weights after backprop
            for name, param in model.named_parameters():
                weights_after_backprop.append(param.clone()) # only layer1's weight should update, layer2 is not used
            
            for i in zip(parameter_name, weights_before_backprop, weights_after_backprop):
                if torch.equal(i[1],i[2]):
                    print(i[0], torch.equal(i[1],i[2]))
            input('halt')
            '''
            '''
            for name, param in model.named_parameters():
                #if not param.requires_grad:
                #    print(name)
                if param.grad!=None:
                    count_nonzero=torch.count_nonzero(param.grad)
                    if count_nonzero.item()!=0:
                        print(name,count_nonzero.item())
                else:
                    print(name,'None')'''
           
            #performance metrics of training dataset
            correct_train,total_images_train,conf_mat_train,_=conf_mat_create(pred, train_labels, correct_train, total_images_train, conf_mat_train)
            print('Train: Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}'.format(epoch+1, epochs, batch_no, batches_train, loss.item()))
        
        if trainingmethod == 'multisteplr1' or trainingmethod == 'lrdecayshu':
            current_lr=scheduler.get_last_lr()[0]
        else:
            current_lr=optimizer.param_groups[0]['lr']
        print("current lr:",current_lr)
        
        running_train_loss=loss_train/total_images_train

        # early_stopping needs the validation loss to check if it has decresed, 
        # and if it has, it will make a checkpoint of the current model
        if patience_epoch:
            correct_test,total_images_test,loss_test,conf_mat_test,auc_val = validation(model, data_iterator_test, epoch, batches_test, class_imbalance)
            valid_loss=loss_test/total_images_test
            results_store_excel(True, True, False, None, correct_train,total_images_train,loss_train,correct_test,total_images_test,loss_test,epoch, conf_mat_train, conf_mat_test, current_lr, auc_val)
            early_stopping(valid_loss,model,optimizer,epoch,conf_mat_train,conf_mat_test, running_train_loss)
            if early_stopping.early_stop:
                print("Early stopping",epoch+1)
                break
        else:
            if use_validation:
                correct_test,total_images_test,loss_test,conf_mat_test,auc_val = validation(model, data_iterator_test, epoch, batches_test, class_imbalance)
                valid_loss=loss_test/total_images_test
                results_store_excel(True, True, False, None, correct_train,total_images_train,loss_train,correct_test,total_images_test,loss_test,epoch, conf_mat_train, conf_mat_test, current_lr, auc_val)
                modelcheckpoint(valid_loss,model,optimizer,epoch,conf_mat_train,conf_mat_test, running_train_loss, auc_val)
            else:
                save_model(model, optimizer, epoch, running_train_loss)
                per_model_metrics, conf_mat_test = test.test(model, dataloader_test, batches_test, activation, sheet2, sheet3, sheet4, device, classes, df_test)
                results_store_excel(True, False, True, per_model_metrics, correct_train,total_images_train,loss_train,None, None, None,epoch, conf_mat_train, None, current_lr)
                sheet3.append(per_model_metrics)
                #correct_test,total_images_test,loss_test,conf_mat_test = validation(model, data_iterator_test, epoch, batches_test, False)

        if trainingmethod == 'multisteplr1' or trainingmethod == 'lrdecayshu': 
            scheduler.step()
    
    if patience_epoch:
        sheet2.append([0,1])
        for row in early_stopping.conf_mat_train_best.tolist():
            sheet2.append(row)
        sheet2.append([0,1])
        for row in early_stopping.conf_mat_test_best.tolist():
            sheet2.append(row)
    elif use_validation:
        sheet2.append([0,1])
        for row in modelcheckpoint.conf_mat_train_best.tolist():
            sheet2.append(row)
        sheet2.append([0,1])
        for row in modelcheckpoint.conf_mat_test_best.tolist():
            sheet2.append(row)
    print('Finished Training')
    
def validation(model, data_iterator_val, epoch, batches_val, class_imbalance1):
    """Validation"""
    model.eval()
    total_images=0
    val_loss = 0
    correct = 0
    s=0
    batch_val_no=0
    conf_mat_test=np.zeros((2,2))
    
    if featureextractormodel=='gmic_resnet18_pretrained':
        if class_imbalance1=='poswt':
            bcelogitloss_val = nn.BCEWithLogitsLoss(pos_weight=class_weights_val[0])
            bceloss_val = nn.BCELoss()
        else:
            bcelogitloss_val = nn.BCEWithLogitsLoss()
            bceloss_val = nn.BCELoss()
    else:
        if class_imbalance1!='focalloss':
            lossfn1 = loss_fn(activation, class_imbalance1, class_weights_val)
    
    with torch.no_grad():   
        for val_idx, val_batch, val_labels in data_iterator_val:
            val_batch, val_labels = val_batch.to(device), val_labels.to(device)
            val_labels = val_labels.view(-1)#.float()
            #with torch.cuda.amp.autocast():
            if featureextractormodel=='gmic_resnet18_pretrained':
                output_batch_local_val, output_batch_global_val, output_batch_fusion_val, saliency_map_val = model(val_batch) # compute model output, loss and total train loss over one epoch
                if class_imbalance1=='poswt':
                    weight_batch_val = torch.tensor([1,class_weights_val[0]]).float().to(device)[val_labels]
                    bceloss_val.weight = weight_batch_val
                output_batch_local_val = output_batch_local_val.view(-1)
                output_batch_global_val = output_batch_global_val.view(-1)
                output_batch_fusion_val = output_batch_fusion_val.view(-1)
                val_labels = val_labels.float()
                val_pred = torch.ge(torch.sigmoid(output_batch_fusion_val), torch.tensor(0.5)).float()
                loss1 = gmic_loss_fn(bcelogitloss_val, bceloss_val, output_batch_local_val, output_batch_global_val, output_batch_fusion_val, saliency_map_val, val_labels, sm_reg_param).item()
                output_val = output_batch_fusion_val
            else:
                output_val = model(val_batch)
                if activation=='sigmoid':
                    if class_imbalance1=='wtcostfunc':
                        weight_batch_val = class_weights_val[val_labels]
                        lossfn1.weight = weight_batch_val
                    output_val = output_val.squeeze(1)
                    output_val = output_val.view(-1)                                                 
                    val_labels=val_labels.float()
                    val_pred = torch.ge(torch.sigmoid(output_val), torch.tensor(0.5)).float()
                    if class_imbalance1=='focalloss':
                        loss1 = sigmoid_focal_loss(output_val, val_labels, alpha=-1, reduction='mean').item()
                    else:
                        loss1 = lossfn1.forward(output_val, val_labels).item()
                elif activation=='softmax':
                    val_pred = output_val.argmax(dim=1, keepdim=True)
                    loss1 = lossfn1(output_val, val_labels).item()
            
            if batch_val_no==0:
                val_pred_all=val_pred
                val_labels_all=val_labels
                print(output_val.data.shape)
                if activation=='sigmoid':
                    output_all_ten=torch.sigmoid(output_val.data)
                elif activation=='softmax':
                    output_all_ten=F.softmax(output_val.data,dim=1)
                    output_all_ten=output_all_ten[:,1]
            else:
                val_pred_all=torch.cat((val_pred_all,val_pred),dim=0)
                val_labels_all=torch.cat((val_labels_all,val_labels),dim=0)
                if activation=='sigmoid':
                    output_all_ten=torch.cat((output_all_ten,torch.sigmoid(output_val.data)),dim=0)
                elif activation=='softmax':
                    output_all_ten=torch.cat((output_all_ten,F.softmax(output_val.data,dim=1)[:,1]),dim=0)

            s=s+val_labels.shape[0]    
            val_loss += val_labels.size()[0]*loss1 # sum up batch loss
            correct,total_images,conf_mat_test,conf_mat_batch=conf_mat_create(val_pred,val_labels,correct,total_images,conf_mat_test)
            
            batch_val_no+=1
            print('Val: Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}'.format(epoch+1, max_epochs, batch_val_no, batches_val, loss1))
    
    print("conf_mat_val:",conf_mat_test)
    print("total_images:",total_images)
    print("s:",s)
    print('\nVal set: total val loss: {:.4f}, Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%), Epoch:{}\n'.format(
        val_loss, val_loss/total_images, correct, total_images,
        100. * correct / total_images,epoch+1))
    
    auc=metrics.roc_auc_score(val_labels_all.cpu().numpy(), output_all_ten.cpu().numpy())
    return correct,total_images,val_loss,conf_mat_test, auc#,loss_ar_val

def data_augmentation(resize):
    # Data augmentation and normalization for training
    # Just normalization for validation
    data_transforms = {
        'train': transforms.Compose([
            utils.MyCrop(),
            transforms.Pad(100),
            transforms.RandomRotation(3),
            transforms.ColorJitter(brightness=0.20, contrast=0.20),
            transforms.RandomAdjustSharpness(sharpness_factor=0.20),
            utils.MyGammaCorrection(0.20),
            utils.MyPaddingLongerSide(resize),
            transforms.Resize((resize[0],resize[1])),
            transforms.ToTensor(),
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ]),
        #'train': transforms.Compose([
        #    transforms.Resize((resize,resize)),
        #    transforms.ToTensor(),
        #    transforms.RandomHorizontalFlip(p=0.5),
        #    transforms.ColorJitter(contrast=0.20, saturation=0.20),
        #    transforms.RandomRotation(30),
        #    AddGaussianNoise(mean=0, std=0.005),
        #    transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        #]),
        'val': transforms.Compose([
            transforms.Resize((resize[0],resize[1])),
            transforms.ToTensor(),
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ]),
    }

    return data_transforms

def seed_worker(worker_id):
    worker_seed = torch.initial_seed() % 2**32
    np.random.seed(worker_seed)
    random.seed(worker_seed)
    
if __name__=='__main__': #before calling train, create the config.ini file for this model run using the code config_file_creation.py
    begin_time = datetime.datetime.now()
    
    #Initialization    
    count=0
    count1=0
    epoch_step=5
    acc_num_list_final=[]
    groundtruth_list=[]
    acc_num_firstsubset=[]
    
    #read arguments
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config_file_path",
        type=str,
        help="full path where the config.ini file containing the parameters to run this code is stored",
    )

    parser.add_argument(
        "--num_config_start",
        type=int,
        help="full path where the config.ini file containing the parameters to run this code is stored",
    )
    
    parser.add_argument(
        "--num_config_end",
        type=int,
        help="full path where the config.ini file containing the parameters to run this code is stored",
    )
    args = parser.parse_args()
    
    num_config_start = args.num_config_start
    num_config_end = args.num_config_end

    '''
    #Read config.ini file
    config_object = ConfigParser()
    with open(args.config_file_path, 'r', encoding='utf-8') as f:
        config_object.read_file(f)
    #config_object.read(args.config_file_path, encoding='utf-8') #this does not work for my code.
    '''

    config_file_names = glob.glob(args.config_file_path+'/config*')
    config_file_names = sorted(config_file_names, key=lambda x: int(re.search(r'\d+$', x.split('.')[-2]).group()))
    print(config_file_names[num_config_start:num_config_end])
    for config_file in config_file_names[num_config_start:num_config_end]:
        print(config_file)
        config_object = ConfigParser()
        with open(config_file, 'r', encoding='utf-8') as f:
            config_object.read_file(f)
    
        #parameters from config.ini for training the model
        rand_seed_other=int(config_object["parametersetting"]['randseedother'])
        rand_seed_data=int(config_object["parametersetting"]['randseeddata'])
        
        #random state initialization of the code - values - 8, 24, 30
        torch.manual_seed(rand_seed_other) 
        torch.cuda.manual_seed(rand_seed_other)
        torch.cuda.manual_seed_all(rand_seed_other)
        np.random.seed(rand_seed_data)
        random.seed(rand_seed_data)
        g = torch.Generator()
        g.manual_seed(rand_seed_other)
        torch.backends.cudnn.deterministic = True
        
        batch_size=int(config_object['parametersetting']['batchsize'])#10
        #print("batch_size:",type(batch_size), batch_size)
        modality=config_object["parametersetting"]['modality']#'MG'
        #print("modality:",type(modality), modality)
        patience_epoch=config_object["parametersetting"]['patienceepoch']#14
        if patience_epoch=='False':
            patience_epoch=False
        else:
            patience_epoch=int(patience_epoch)
        #print("patience_epoch:",type(patience_epoch), patience_epoch)
        num_classes=int(config_object["parametersetting"]['numclasses'])
        #print("num_classes:",type(num_classes), num_classes)
        max_epochs=int(config_object["parametersetting"]['maxepochs'])
        #max_epochs=2
        #print("max_epochs:",type(max_epochs), max_epochs)
        num_workers=int(config_object["parametersetting"]['numworkers'])
        #print("num_workers:",type(num_workers), num_workers)
        groundtruth_dic=ast.literal_eval(config_object["parametersetting"]['groundtruthdic'])
        #print("groundtruth_dic:",type(groundtruth_dic),groundtruth_dic)
        classes=ast.literal_eval(config_object["parametersetting"]['classes'])
        #print("classes:",type(classes), classes)
        resize=ast.literal_eval(config_object["parametersetting"]['resize'])
        #print("resize:",type(resize), resize)
        activation=config_object["parametersetting"]['activation']#'sigmoid' #Kim:softmax; MaxWelling:sigmoid; Kim_maxpool:softmax
        #print("activation:",type(activation), activation)
        class_imbalance=config_object["parametersetting"]['classimbalance']
        if class_imbalance=='False':
            class_imbalance=False
        #print("class_imbalance:",type(class_imbalance), class_imbalance)
        milpooling=config_object["parametersetting"]['milpooling']#'attention' #Kim:average; MaxWelling:attention; Kim_maxpool: maxpool
        #print("milpooling:",type(milpooling), milpooling)
        attention=config_object["parametersetting"]['attention']#'separate_pipeline' #separate pipeline model
        if attention=='False':
            attention=False
        #print("attention:",type(attention), attention)
        feature_extractor=config_object["parametersetting"]['featureextractor']#'separate_pipeline' #separate pipeline model
        #print("feature_extractor:",type(feature_extractor), feature_extractor)
        data_type=config_object["parametersetting"]['data']
        #data_type='fixed'
        #print("data:",type(data_type), data_type)
        data_aug=config_object["parametersetting"]['dataaug']
        #data_aug='big'
        dataset=config_object["parametersetting"]['dataset']
        #print("data_aug:",type(data_aug), data_aug)
        baseline=config_object["parametersetting"]['baseline']
        if baseline=='False':
            baseline=False
        else:
            baseline==True
        datascaling=config_object["parametersetting"]['datascaling']
        if datascaling=='False':
            datascaling=False
        extra=config_object["parametersetting"]['extra']
        if extra=='False':
            extra=False
        flipimage=config_object["parametersetting"]['flipimage']
        if flipimage=='False':
            flipimage=False
        else:
            flipimage==True
        featurenorm=config_object["parametersetting"]['featurenorm']
        if featurenorm=='False':
            featurenorm=False
        if featurenorm=='rgb':
            inchans=3
        else:
            inchans=1
        featureextractormodel=config_object["parametersetting"]['femodel']
        if featureextractormodel=='False':
            featureextractormodel=False
        trainingmethod=config_object["parametersetting"]['trainingmethod']
        if trainingmethod=='False':
            trainingmethod=False
        optimizer_type=config_object["parametersetting"]['optimizer']
        
        try:
            run_descrip=config_object["parametersetting"]['run']
            if run_descrip=='False':
                run_descrip=False
        except:
            run_descrip=False
        try:
            lrval=float(config_object["parametersetting"]['lr'])
        except:
            lrval=False
        try:
            wtdecay=float(config_object["parametersetting"]['wtdecay'])
        except:
            wtdecay=False
        try:
            topkpatch=config_object["parametersetting"]['topkpatch']
            if topkpatch=='False':
                topkpatch=False
            else:
                topkpatch=float(topkpatch)
        except:
            topkpatch=False   
        
        try:
            use_validation=config_object["parametersetting"]['use_validation']
            if use_validation=='False':
                use_validation=False
            else:
                use_validation=True
        except:
            use_validation=='False'

        try:
            sm_reg_param = config_object["parametersetting"]['sm_reg_param']
            if sm_reg_param=='False':
                sm_reg_param=False
            else:
                sm_reg_param=float(sm_reg_param)
        except:
            sm_reg_param=False

        try:
            image_cleaning = config_object["parametersetting"]['image_cleaning']
            if image_cleaning=='False':
                image_cleaning=False
        except:
            image_cleaning=False
        
        #GMIC
        gmic_parameters = {
            "device_type": 'gpu',
            "gpu_number": str(config_object["parametersetting"]['device'].split(':')[1]),
            "max_crop_noise": (100, 100),
            "max_crop_size_noise": 100,
            # model related hyper-parameters
            "cam_size": (92, 60),
            "K": 6,
            "crop_shape": (256, 256),
            "post_processing_dim":512,
            "num_classes":1,
            "use_v1_global":True,
            "percent_t": topkpatch,
            'arch':'resnet18',
            'pretrained': True
        }
        
        # CUDA for PyTorch
        use_cuda = torch.cuda.is_available()
        device = torch.device(config_object["parametersetting"]['device'] if use_cuda else "cpu")
        #device = torch.device('cuda:2' if use_cuda else "cpu")
        print(device)
        
        #input file names
        #new addition
        if extra=='IL' or extra=='CL':
            csv_file_modality='/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_multiinstance_groundtruth.csv'
        else:
            csv_file_modality='/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_singleinstance_groundtruth.csv' # I changed this inputto single-instance input
        df_modality=pd.read_csv(csv_file_modality,sep=';')
        #df_modality=df_modality[~df_modality['StudyInstanceUID'].isnull()]
        print("the original df modality shape:",df_modality.shape)
        df_modality=df_modality[~df_modality['Views'].isnull()]
        print("df modality no null view:",df_modality.shape)
        #df_modality = df_modality[:200]
        #utils.crosscheck_view_collect_images(df_modality)
        #w_mean, w_std, h_mean, h_std = utils.calculate_image_size(df_modality)
        #print(w_mean, w_std, h_mean, h_std)
        #input('halt')
        
        #View distribution and creates a new file with the view names mentioned and saves in input_data
        #utils.views_distribution(df_modality)
        #utils.plot('views_dic_allowed.xlsx')
        #input('wait')
        
        #output files path
        #file_name=args.config_file_path.split('/')[-2]
        file_name=config_file.split('/')[-2]
        print(file_name)
        
        path_to_output="/homes/spathak/multiview_mammogram/models_results/cbis-ddsm/newstoryline1/"+file_name+"/"
        
        if rand_seed_data!=rand_seed_other:
            rand_seed = str(rand_seed_other) +'_'+ str(rand_seed_data)
        else:
            rand_seed = rand_seed_data
        
        path_to_hyperparam_search = path_to_output+"hyperparamsearch1_"+str(num_config_start)+'-'+str(num_config_end)+'_'+str(rand_seed)+".xlsx"
        
        if run_descrip:
            path_to_model=path_to_output+"model1_"+str(rand_seed)+'_'+run_descrip+".tar"
            path_to_results=path_to_output+"result1_"+str(rand_seed)+'_'+run_descrip+".xlsx"
            path_to_results_text=path_to_output+"result1_"+str(rand_seed)+'_'+run_descrip+".txt"
            path_to_learning_curve=path_to_output+"learningcurve1_"+str(rand_seed)+'_'+run_descrip+".png"
            path_to_log_file=path_to_output+"log1_"+str(rand_seed)+'_'+run_descrip+".txt"

        else:
            path_to_model=path_to_output+"model1_"+str(rand_seed)+".tar"
            path_to_results=path_to_output+"result1_"+str(rand_seed)+".xlsx"
            path_to_results_text=path_to_output+"result1_"+str(rand_seed)+".txt"
            path_to_learning_curve=path_to_output+"learningcurve1_"+str(rand_seed)+".png"
            path_to_log_file=path_to_output+"log1_"+str(rand_seed)+".txt"
        
        #check output_folder path
        if not os.path.exists(path_to_output):
            print("Error! config file path does not exist! This code needs the same path to store the output files and model.")
            sys.exit()
        
        #new data file after adding view information to the file and taking instances with exactly 4 views
        #df_modality=df_modality[:400]
        #df_modality1=df_modality[df_modality['Views'].str.split('+').str.len()==4.0] #I only changed this here for adapting to single-instance
        #print("df_modality 4 views:", df_modality1.shape)
        
        #Train-val-test split
        if dataset=='officialtestset':
            #if 16 bit png
            #df_modality['FullPath']=df_modality['FullPath'].str.rsplit('/',n=2).str[0]+'/'+'processed-images-gmiccleaningcode'+'/'+df_modality['FullPath'].str.rsplit('/',n=1).str[-1]
            df_modality['FullPath']=df_modality['FullPath'].str.rsplit('/',n=2).str[0]+'/'+'processed-images-gmic'+'/'+df_modality['FullPath'].str.rsplit('/',n=1).str[-1]
            df_train = df_modality[df_modality['ImageName'].str.contains('Training')]
            if patience_epoch or use_validation:
                df_train, df_val = train_test_split(df_train,test_size=0.10,shuffle=True,stratify=df_train['Groundtruth'])
            df_test = df_modality[df_modality['ImageName'].str.contains('Test')]
            #df_train = df_train[:20]
            #df_val = df_val[:5]
            #df_test = df_test[:10]
            #print(df_train)
            #print(df_val)
            #print(df_test)
            total_instances=df_modality.shape[0]
        else:
            #new addition
            #new data file after adding view information to the file and taking instances with exactly 4 views
            #df_modality=df_modality[:400]
            df_modality1=df_modality[df_modality['Views'].str.split('+').str.len()==4.0]
            print("df_modality 4 views:", df_modality1.shape)
            df_train, df_val, df_test = utils.stratifiedgroupsplit(df_modality1, rand_seed_data)
            total_instances=df_modality1.shape[0]
            #if data_type=='variable':
            #bags with views<4
            df_modality2=df_modality[df_modality['Views'].str.split('+').str.len()!=4.0]
            print("df_modality views<4:",df_modality2.shape)
        
            df_train = pd.concat([df_train, df_modality2[df_modality2['Patient_Id'].isin(df_train['Patient_Id'].unique().tolist())]])
            df_val = pd.concat([df_val, df_modality2[df_modality2['Patient_Id'].isin(df_val['Patient_Id'].unique().tolist())]])
            df_test = pd.concat([df_test, df_modality2[df_modality2['Patient_Id'].isin(df_test['Patient_Id'].unique().tolist())]])
            df_modality2 = df_modality2[~df_modality2['Patient_Id'].isin(df_train['Patient_Id'].unique().tolist()+df_val['Patient_Id'].unique().tolist()+df_test['Patient_Id'].unique().tolist())]
            
            df_train1, df_val1, df_test1 = utils.stratifiedgroupsplit(df_modality2, rand_seed_data)
            
            df_train=pd.concat([df_train,df_train1])
            df_val=pd.concat([df_val,df_val1])
            df_test=pd.concat([df_test,df_test1])
            
            total_instances=df_modality.shape[0]
            print("Check starting between perfect transfer of patients from case based to single instance based")
            train_check=df_train['FolderName'].unique().tolist()
            val_check=df_val['FolderName'].unique().tolist()
            test_check=df_test['FolderName'].unique().tolist()
            train_check.sort()
            val_check.sort()
            test_check.sort()
            print(len(train_check))
            print(len(val_check))
            print(len(test_check))
        
        #print(df_train.index)
        #print(df_val.index)
        #print(df_test.index)
        
        
        
        #new addition
        if extra=='IL':
            df_instances=pd.read_csv('/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_singleinstance_groundtruth.csv', sep=';')
            df_train = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_train['FolderName'].tolist())]
            df_val = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_val['FolderName'].tolist())]
            df_test = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_test['FolderName'].tolist())]
        
        if extra=='CL':
            df_instances=pd.read_csv('/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_singleinstance_caselabel_groundtruth.csv', sep=';')
            '''
            df_instances=pd.read_csv('/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_singleinstance_groundtruth.csv', sep=';')
            df_instances['FolderName']=df_instances['ImageName'].str.split('_').str[:3].str.join('_')
            trial = df_instances.groupby(by='FolderName').apply(lambda x: len(x['Groundtruth'].unique())>1)
            print(trial[trial])
            for index in df_instances.index:
                before_gt=df_instances.loc[index,'Groundtruth']
                df_instances.loc[index,'Groundtruth']=df_modality[df_modality['FolderName']=='_'.join(df_instances.loc[index,'ImageName'].split('_')[:3])]['Groundtruth'].item()
                if before_gt!=df_instances.loc[index,'Groundtruth']:
                    print("before, after:",df_instances.loc[index,'Patient_Id'], before_gt, df_instances.loc[index,'Groundtruth'])
            df_instances.to_csv('/projects/dso_mammovit/project_kushal/data/MG_training_files_cbis-ddsm_singleinstance_caselabel_groundtruth.csv',sep=';',na_rep='NULL', index=False)
            '''
            df_train = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_train['FolderName'].tolist())]
            df_val = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_val['FolderName'].tolist())]
            df_test = df_instances[df_instances['ImageName'].str.split('_').str[:3].str.join('_').isin(df_test['FolderName'].tolist())]
        
        if extra=='IL' or extra=='CL':
            train_check1=df_train['ImageName'].str.split('_').str[:3].str.join('_').unique().tolist()
            val_check1=df_val['ImageName'].str.split('_').str[:3].str.join('_').unique().tolist()
            test_check1=df_test['ImageName'].str.split('_').str[:3].str.join('_').unique().tolist()
            train_check1.sort()
            val_check1.sort()
            test_check1.sort()
            print(len(train_check1))
            print(len(val_check1))
            print(len(test_check1))
            if train_check==train_check1:
                print('identical train')
            if val_check==val_check1:
                print('identical val')
            if test_check==test_check1:
                print('identical test')
            
            print("check end!")
        
        df_train = df_train.reset_index()
        if patience_epoch or use_validation:
            df_val = df_val.reset_index()
        df_test = df_test.reset_index()
        
        train_instances=df_train.shape[0]
        print("Train:",utils.stratified_class_count(df_train))
        print("training instances:", train_instances)
        
        if patience_epoch or use_validation:
            val_instances=df_val.shape[0]
            print("Val:",utils.stratified_class_count(df_val))
            print("Validation instances:", val_instances)
        
        test_instances=df_test.shape[0]
        print("Test:",utils.stratified_class_count(df_test)) 
        print("Test instances:",test_instances)
        
        if data_type=='variable':
            #group by view
            df_train, view_group_names_train = utils.groupby_view(df_train)
            print(view_group_names_train)
            
            df_val, view_group_names_val = utils.groupby_view(df_val)
            print(view_group_names_val)
            
            df_test, view_group_names_test = utils.groupby_view(df_test)
            print(view_group_names_test)
        
        if class_imbalance=='wtcostfunc':
            class_weights_train = utils.class_distribution_weightedloss(df_train)
            class_weights_val = utils.class_distribution_weightedloss(df_val)
            class_weights_train, class_weights_val = class_weights_train.to(device), class_weights_val.to(device)
        elif class_imbalance=='poswt':
            class_weights_train = utils.class_distribution_poswt(df_train)
            class_weights_train = class_weights_train.to(device)
            if patience_epoch or use_validation:
                class_weights_val = utils.class_distribution_poswt(df_val)
                class_weights_val = class_weights_val.to(device) 
            else:
                class_weights_val = None     
        else:
            class_weights_train = None
            class_weights_val = None
        
        
        # set file path
        if os.path.isfile(path_to_results):
            wb = op.load_workbook(path_to_results)
            sheet1 = wb['train_val_loss_acc']
            sheet2 = wb['confusion_matrix_train_val']
            sheet3 = wb['confusion_matrix_test']
            sheet4 = wb['metrics_view_wise']
        else:
            wb=Workbook()
            sheet1 = wb.active
            sheet1.title="train_val_loss_acc"
            header=['Epoch','lr','Avg Loss Train','Accuracy Train','F1macro Train','Recall Train','Speci Train','Avg Loss Val','Accuracy Val','F1macro Val','Recall Val','Speci Val','AUC Val']
            sheet1.append(header)
            sheet2 = wb.create_sheet('confusion_matrix_train_val')
            sheet3 = wb.create_sheet('confusion_matrix_test') 
            sheet4 = wb.create_sheet('metrics_view_wise')
        
        # set file path
        if os.path.isfile(path_to_hyperparam_search):
            wb1 = op.load_workbook(path_to_hyperparam_search)
            sheet5 = wb1['hyperparam_results']
        else:
            wb1=Workbook()
            sheet5 = wb1.active
            sheet5.title="hyperparam_results"
            header=['config_file','lr','wtdecay','sm_reg_param','trainingscheme','optimizer','patienceepoch','batchsize','Loss','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC']
            sheet5.append(header)
        
        #model = models.resnet18(pretrained=True)
        #num_ftrs = model.fc.in_features
        #model.fc = nn.Linear(num_ftrs, 1)
        print(activation, featureextractormodel, extra)
        model = mymodels.SILmodel(activation, featureextractormodel, extra, topkpatch, gmic_parameters)
        #model = model.SeparatePipelineMIL(milpooling, activation, device, extra, attention, feature_extractor)
        #print(model)
        for name, param in model.named_parameters():
            if param.requires_grad:
                print(name)
        #        #if 'fc' in name:
        #        #    print(param)
        #input('halt')
        #test_input=torch.randn((4,1,1600,1600))
        #summary(model,test_input,[4],['LCC','LMLO','RCC','RMLO'],depth=5)
        model.to(device)
        pytorch_total_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
        print(pytorch_total_params)
        
        #image standardization
        if datascaling=='scaling':
            if featureextractormodel:
                if 'resnet18pretrained' in featureextractormodel or 'resnet34pretrained' in featureextractormodel or 'resnet50pretrained' in featureextractormodel or 'densenet' in featureextractormodel or 'gmic_resnet18_pretrained' in featureextractormodel or 'convnext-T' in featureextractormodel:
                    if featurenorm=='rgb' or featurenorm=='rchan':
                        mean=[0.485, 0.456, 0.406]
                        std_dev=[0.229, 0.224, 0.225]
                    elif featurenorm=='avg':
                        mean=[0.485, 0.456, 0.406]
                        std_dev=[0.229, 0.224, 0.225]
                        mean_avg=sum(mean)/len(mean)
                        std_dev_avg=sum(std_dev)/len(std_dev)
                        mean = [mean_avg,mean_avg,mean_avg]
                        std_dev = [std_dev_avg,std_dev_avg,std_dev_avg]
                    elif featurenorm==False:
                        mean=[0.5,0.5,0.5]
                        std_dev=[0.5,0.5,0.5]
                else:
                    mean=[0.5,0.5,0.5]
                    std_dev=[0.5,0.5,0.5]
            else:
                mean=[0.5,0.5,0.5]
                std_dev=[0.5,0.5,0.5]
        elif datascaling=='standardize':
            #mean, std_dev = utils.calculate_dataset_mean_stddev(df_train, resize, transform=True)
            mean = [0.2016,0.2016,0.2016]
            std_dev = [0.1953,0.1953,0.1953]
            print("mean, std dev:",mean,std_dev)
        else:
            mean=None
            std_dev=None
        
        print(mean)
        print(std_dev)
        if data_aug=='small':
            #print("baseline")
            preprocess_train = transforms.Compose([
                transforms.ColorJitter(brightness=0.10, contrast=0.10),
                transforms.Resize((resize[0],resize[1])),
                transforms.ToTensor(),
                transforms.Normalize(mean=mean, std=std_dev)
            ])
            preprocess_val=utils.data_augmentation_test(mean,std_dev,resize,datascaling)
        
        elif data_aug=='gmic':
            preprocess_train = utils.data_augmentation_train_shen_gmic(mean,std_dev,resize,datascaling,image_cleaning)
            preprocess_val = utils.data_augmentation_test_shen_gmic(mean,std_dev,resize,datascaling,image_cleaning)
        
        elif data_aug=='shu':
            preprocess_train=utils.data_augmentation_train_shu(mean,std_dev,resize,datascaling)
            preprocess_val=utils.data_augmentation_test(mean,std_dev,resize,datascaling)
        
        else:
            preprocess_train=utils.data_augmentation_train(mean,std_dev,resize,datascaling)
            preprocess_val=utils.data_augmentation_test(mean,std_dev,resize,datascaling)
        
        
        batch_sampler=None
        batch_sampler_val=None
        batch_sampler_test=None
        shuffle=True
        sampler1=None
        if data_type=='variable':
            if class_imbalance=='oversampling':
                sampler = utils.CustomGroupbyViewWeightedRandomSampler(df_train)
                sampler_val = utils.CustomGroupbyViewRandomSampler(df_val)
            else:
                sampler = utils.CustomGroupbyViewRandomSampler(df_train)
                sampler_val = utils.CustomGroupbyViewRandomSampler(df_val)
            
            sampler_test = utils.CustomGroupbyViewRandomSampler(df_test)
            
            view_group_length, view_group_name = sampler.__viewlength__()
            view_group_length_val, view_group_name_val = sampler_val.__viewlength__()
            view_group_length_test, view_group_name_test = sampler_test.__viewlength__()
            batch_sampler = utils.CustomGroupbyViewRandomBatchSampler(sampler, batch_size, view_group_length, view_group_name)
            batch_sampler_val = utils.CustomGroupbyViewRandomBatchSampler(sampler_val, batch_size, view_group_length_val, view_group_name_val)
            batch_sampler_test = utils.CustomGroupbyViewRandomBatchSampler(sampler_test, batch_size, view_group_length_test, view_group_name_test)
            batch_size1=1
            shuffle=False
        else: 
            if class_imbalance=='oversampling':
                sampler1 = utils.CustomWeightedRandomSampler(df_train)
                shuffle = False
            batch_size1=batch_size
        
        dataset_gen_train = utils.BreastCancerDataset_generator(df_train,modality, datascaling, resize, flipimage, inchans, image_cleaning, preprocess_train)
        dataloader_train = DataLoader(dataset_gen_train, batch_size=batch_size1, shuffle=shuffle, num_workers=num_workers, collate_fn=utils.MyCollate, worker_init_fn=seed_worker, generator=g, sampler=sampler1, batch_sampler=batch_sampler)    
        
        if patience_epoch or use_validation:
            dataset_gen_val = utils.BreastCancerDataset_generator(df_val, modality, datascaling, resize, flipimage, inchans, image_cleaning, preprocess_val)
            dataloader_val = DataLoader(dataset_gen_val, batch_size=batch_size1, shuffle=False, num_workers=num_workers, collate_fn=utils.MyCollate, worker_init_fn=seed_worker, generator=g, batch_sampler=batch_sampler_val)
            #worker_init_fn=seed_worker, generator=g,
        
        dataset_gen_test = utils.BreastCancerDataset_generator(df_test, modality, datascaling, resize, flipimage, inchans, image_cleaning, preprocess_val)
        dataloader_test = DataLoader(dataset_gen_test, batch_size=batch_size1, shuffle=False, num_workers=num_workers, collate_fn=utils.MyCollate, worker_init_fn=seed_worker, generator=g, batch_sampler=batch_sampler_test)
        
        #writer = SummaryWriter('/homes/spathak/multiview_mammogram/runs/'+file_name)
        #writer=tensorboard_log.tensorboard_log(max_epochs, file_name, dataloader_train, model)
        
        if data_type=='variable':
            batches_train=int(sum(np.ceil(np.array(list(view_group_names_train.values()))/batch_size)))
            batches_val=int(sum(np.ceil(np.array(list(view_group_names_val.values()))/batch_size)))
            batches_test=int(sum(np.ceil(np.array(list(view_group_names_test.values()))/batch_size)))
            #print(batches_train)
            #print(batches_val)
        else:
            if class_imbalance=='oversampling':
                batches_train=int(math.ceil(sampler1.__len__()/batch_size))
            else:
                batches_train=int(math.ceil(train_instances/batch_size))
            
            if patience_epoch or use_validation:
                batches_val=int(math.ceil(val_instances/batch_size))
            batches_test=int(math.ceil(test_instances/batch_size))
        
        
        #training the model
        if patience_epoch or use_validation:
            train(model, dataloader_train, dataloader_val, batches_train, batches_val, max_epochs)
        else:
            sheet3.append(['Loss','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC'])
            train(model, dataloader_train, dataloader_test, batches_train, batches_test, max_epochs)

        #testing the model
        path_to_trained_model = path_to_model
        model1 = load_model_for_testing(model, path_to_trained_model)
        per_model_metrics_val, _ = test.test(model1, dataloader_val, batches_val, activation, sheet2, sheet3, sheet4, device, classes, df_val, featureextractormodel, sm_reg_param)
        sheet3.append(['Loss','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC'])
        per_model_metrics_test, conf_mat_test = test.test(model1, dataloader_test, batches_test, activation, sheet2, sheet3, sheet4, device, classes, df_test, featureextractormodel, sm_reg_param)
        
        sheet2.append([0,1])
        for row in conf_mat_test.tolist():
            sheet2.append(row)
            
        sheet3.append(per_model_metrics_test)
        #sheet3, sheet4 = test.run_test(df_test, modality, preprocess_val, batch_size1, num_workers, batches_test, activation, sheet3, sheet4, path_to_model, device, feature_extractor, attention, classes, milpooling, extra, datascaling, resize, attention, batch_sampler_test, flipimage, inchans)

        #save the results
        wb.save(path_to_results)

        #plot the training and validation loss and accuracy
        #df=pd.read_excel(path_to_results)
        #results_plot(df,file_name)
        
        hyperparam_details = [config_file.split('/')[-1], lrval, wtdecay, sm_reg_param, trainingmethod, optimizer_type, patience_epoch, batch_size] + per_model_metrics_val
        sheet5.append(hyperparam_details)
        wb1.save(path_to_hyperparam_search)

        '''except Exception as err:
            print("Exception encountered:",err)
            #save the results
            wb.save(path_to_results)
        
            #plot the training and validation loss and accuracy
            #df=pd.read_excel(path_to_results)
            #results_plot(df,file_name)
        '''
        f = open(path_to_log_file,'w')
        f.write("Model parameters:"+str(pytorch_total_params/math.pow(10,6))+'\n')
        f.write("Start time:"+str(begin_time)+'\n')
        f.write("End time:"+str(datetime.datetime.now())+'\n')
        f.write("Execution time:"+str(datetime.datetime.now() - begin_time)+'\n')
        f.close()
    
