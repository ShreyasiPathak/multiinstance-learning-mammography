import torch
import itertools
import numpy as np
import openpyxl as op
import matplotlib.pyplot as plt
import torch.nn.functional as F
from sklearn.metrics import confusion_matrix

from utilities import utils

def results_store_excel(train_res, val_res, test_res, per_model_metrics, correct_train, total_images_train, train_loss, correct_test, total_images_test, test_loss, epoch, conf_mat_train, conf_mat_test, lr, auc_val, path_to_results, path_to_results_text):
    lines = [epoch+1, lr]
    if train_res:
        avg_train_loss=train_loss/total_images_train
        accuracy_train=correct_train / total_images_train
        speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
        prec_train=conf_mat_train[1,1]/sum(conf_mat_train[:,1])
        f1_train=2*recall_train*prec_train/(recall_train+prec_train)
        prec_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[:,0])
        recall_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        f1_train_neg=2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
        f1macro_train=(f1_train+f1_train_neg)/2
        lines.extend([avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train])

    if val_res:
        speci_test=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        avg_test_loss=test_loss/total_images_test
        recall_test=conf_mat_test[1,1]/sum(conf_mat_test[1,:])
        prec_test=conf_mat_test[1,1]/sum(conf_mat_test[:,1])
        f1_test=2*recall_test*prec_test/(recall_test+prec_test)
        accuracy_test=correct_test / total_images_test
        recall_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        prec_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[:,0])
        f1_test_neg=2*recall_test_neg*prec_test_neg/(recall_test_neg+prec_test_neg)
        f1macro_test=(f1_test+f1_test_neg)/2
        lines.extend([avg_test_loss, accuracy_test, f1macro_test, recall_test, speci_test, auc_val])

    if test_res:
        lines.extend(per_model_metrics)
    
    out = open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    out.close()
    write_results_xlsx(lines, path_to_results, 'train_val_results')

def results_plot(df, file_name):
    plt.plot(df['Epoch'],df['F1macro Train'],'-r',label='Train F1macro')
    plt.plot(df['Epoch'],df['F1macro Val'],'-b',label='Val F1macro')
    plt.plot(df['Epoch'],df['Avg Loss Train'],'-g',label='Train Loss')
    plt.plot(df['Epoch'],df['Avg Loss Val'],'-y',label='Val Loss')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,df.iloc[-1]['Epoch']))
    plt.xlabel('Epochs')
    plt.ylabel('Loss/F1macro')
    plt.title('Learning curve')
    plt.savefig(file_name, bbox_inches='tight')

def conf_mat_create(predicted, true, correct, total_images, conf_mat, classes):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images, conf_mat, conf_mat_batch

def write_results_xlsx(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)

def write_results_xlsx_confmat(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append([0,1])
    for row in results.tolist():
        sheet.append(row)
    wb.save(path_to_results)

def calc_viewwise_metric(views_names, views_standard, count_dic_viewwise, test_labels, test_pred, output_test):
    views_names_key="+".join(views_names)
    views_count_key=str(len(views_names))

    flag=0
    for view in views_names:
        if view not in views_standard:
            flag=1
    if flag==0:
        views_count_4std_key='views_std'
        if views_count_4std_key in count_dic_viewwise.keys():
            count_dic_viewwise[views_count_4std_key]['true'] = np.append(count_dic_viewwise[views_count_4std_key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[views_count_4std_key]['pred'] = np.append(count_dic_viewwise[views_count_4std_key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[views_count_4std_key]['prob'] = np.append(count_dic_viewwise[views_count_4std_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[views_count_4std_key]={}
            count_dic_viewwise[views_count_4std_key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[views_count_4std_key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[views_count_4std_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    elif flag==1:
        views_count_4nonstd_key='views_nonstd'
        if views_count_4nonstd_key in count_dic_viewwise.keys():
            count_dic_viewwise[views_count_4nonstd_key]['true'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[views_count_4nonstd_key]['pred'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[views_count_4nonstd_key]['prob'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[views_count_4nonstd_key]={}
            count_dic_viewwise[views_count_4nonstd_key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[views_count_4nonstd_key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[views_count_4nonstd_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    if views_names_key in count_dic_viewwise.keys():
        count_dic_viewwise[views_names_key]['true'] = np.append(count_dic_viewwise[views_names_key]['true'],test_labels.cpu().numpy())
        count_dic_viewwise[views_names_key]['pred'] = np.append(count_dic_viewwise[views_names_key]['pred'],test_pred.cpu().numpy())
        count_dic_viewwise[views_names_key]['prob'] = np.append(count_dic_viewwise[views_names_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
    else:
        count_dic_viewwise[views_names_key]={}
        count_dic_viewwise[views_names_key]['true'] = test_labels.cpu().numpy()
        count_dic_viewwise[views_names_key]['pred'] = test_pred.cpu().numpy()
        count_dic_viewwise[views_names_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    
    if views_count_key in count_dic_viewwise.keys():
        count_dic_viewwise[views_count_key]['true'] = np.append(count_dic_viewwise[views_count_key]['true'],test_labels.cpu().numpy())
        count_dic_viewwise[views_count_key]['pred'] = np.append(count_dic_viewwise[views_count_key]['pred'],test_pred.cpu().numpy())
        count_dic_viewwise[views_count_key]['prob'] = np.append(count_dic_viewwise[views_count_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
    else:
        count_dic_viewwise[views_count_key]={}
        count_dic_viewwise[views_count_key]['true'] = test_labels.cpu().numpy()
        count_dic_viewwise[views_count_key]['pred'] = test_pred.cpu().numpy()
        count_dic_viewwise[views_count_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    return count_dic_viewwise

def write_results_viewwise(path_to_results_xlsx, sheetname, count_dic_viewwise):
    wb = op.load_workbook(path_to_results_xlsx)
    if count_dic_viewwise!={}:
        for key in count_dic_viewwise.keys():
            print(key)
            per_model_metrics = utils.performance_metrics(None,count_dic_viewwise[key]['true'],count_dic_viewwise[key]['pred'],count_dic_viewwise[key]['prob'])
            header = [key]
            sheet = wb[sheetname]
            sheet.append(header)
            sheet.append(['Count','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC'])
            sheet.append([len(count_dic_viewwise[key]['true'])]+per_model_metrics)
    wb.save(path_to_results_xlsx)

def case_label_from_SIL(config_params, df_test, test_labels_all, test_pred_all):
    dic_true={}
    dic_pred={}
    idx=0
    
    if config_params['dataset'] == 'mgm':
        image_col = 'CasePath'
    elif config_params['dataset'] == 'cbis-ddsm':
        image_col = 'ImageName'
    elif config_params['dataset'] == 'vindr':
        image_col = 'StudyInstanceUID'


    test_pred_all = test_pred_all.reshape(-1)
    #print(test_labels_all)
    #print(test_pred_all)
    for idx in df_test.index:
        if config_params['dataset'] == 'cbis-ddsm':
            dic_key = '_'.join(df_test.loc[idx, image_col].split('_')[:3])
        elif config_params['dataset'] == 'mgm':
            dic_key = df_test.loc[idx, image_col].split('/')[-1]
        else:
            dic_key = df_test.loc[idx, image_col]
        dic_true[dic_key] = max(dic_true.get(dic_key,0), test_labels_all[idx])
        dic_pred[dic_key] = max(dic_pred.get(dic_key,0), test_pred_all[idx])
    case_labels_true = np.array(list(dic_true.values()))
    case_labels_pred = np.array(list(dic_pred.values()))

    #print(case_labels_true)
    #print(case_labels_pred)
    metrics_case_labels = utils.performance_metrics(None, case_labels_true, case_labels_pred, None)
    print("case label:", metrics_case_labels)
    #sheet3.append(['Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC'])
    #sheet3.append(metrics_case_labels)
    return metrics_case_labels
