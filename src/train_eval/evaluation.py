import torch
import numpy as np
import openpyxl as op
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn import metrics
#import mlflow

def results_store_excel(train_res, val_res, test_res, per_model_metrics, correct_train, total_images_train, train_loss, correct_test, total_images_test, test_loss, epoch, conf_mat_train, conf_mat_test, lr, auc_val, auc_valmacro, path_to_results, path_to_results_text):
    lines = [epoch+1, lr]
    if train_res:
        avg_train_loss=train_loss/total_images_train
        accuracy_train=correct_train / total_images_train
        speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
        prec_train=conf_mat_train[1,1]/sum(conf_mat_train[:,1])
        f1_train=2*recall_train*prec_train/(recall_train+prec_train)
        prec_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[:,0])
        recall_train_neg=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
        f1_train_neg=2*recall_train_neg*prec_train_neg/(recall_train_neg+prec_train_neg)
        f1macro_train=(f1_train+f1_train_neg)/2
        lines.extend([avg_train_loss, accuracy_train, f1macro_train, recall_train, speci_train])
        metrics_dic_train = {"loss_train": avg_train_loss, "accuracy_train": accuracy_train, "f1macro_train": f1macro_train}
        #mlflow.log_metrics(metrics_dic_train, step=epoch)

    if val_res:
        speci_test=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        avg_test_loss=test_loss/total_images_test
        recall_test=conf_mat_test[1,1]/sum(conf_mat_test[1,:])
        prec_test=conf_mat_test[1,1]/sum(conf_mat_test[:,1])
        f1_test=2*recall_test*prec_test/(recall_test+prec_test)
        accuracy_test=correct_test / total_images_test
        recall_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
        prec_test_neg=conf_mat_test[0,0]/sum(conf_mat_test[:,0])
        f1_test_neg=2*recall_test_neg*prec_test_neg/(recall_test_neg+prec_test_neg)
        f1macro_test=(f1_test+f1_test_neg)/2
        lines.extend([avg_test_loss, accuracy_test, f1macro_test, recall_test, speci_test, auc_val, auc_valmacro])
        metrics_dic_val = {"loss_val": avg_test_loss, "accuracy_val": accuracy_test, "f1macro_val": f1macro_test} 
        #mlflow.log_metrics(metrics_dic_val, step=epoch)      

    if test_res:
        lines.extend(per_model_metrics)
    
    out = open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    out.close()
    write_results_xlsx(lines, path_to_results, 'train_val_results')

def results_plot(df, file_name):
    plt.plot(df['Epoch'],df['F1macro Train'],'-r',label='Train F1macro')
    plt.plot(df['Epoch'],df['F1macro Val'],'-b',label='Val F1macro')
    plt.plot(df['Epoch'],df['Avg Loss Train'],'-g',label='Train Loss')
    plt.plot(df['Epoch'],df['Avg Loss Val'],'-y',label='Val Loss')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,df.iloc[-1]['Epoch']))
    plt.xlabel('Epochs')
    plt.ylabel('Loss/F1macro')
    plt.title('Learning curve')
    plt.savefig(file_name, bbox_inches='tight')

def conf_mat_create(predicted, true, correct, total_images, conf_mat, classes):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images, conf_mat, conf_mat_batch

def write_results_xlsx(results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    sheet = wb[sheetname]
    sheet.append(results)
    wb.save(path_to_results)

def write_results_xlsx_confmat(config_params, results, path_to_results, sheetname):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    
    sheet.append(config_params['classes'])
    for row in results.tolist():
        sheet.append(row)
    wb.save(path_to_results)

def calc_viewwise_metric(views_names, views_standard, count_dic_viewwise, test_labels, test_pred, output_test):
    views_names_key="+".join(views_names)
    views_count_key=str(len(views_names))

    flag=0
    for view in views_names:
        if view not in views_standard:
            flag=1
    if flag==0:
        views_count_std_key='views_std'
        if views_count_std_key in count_dic_viewwise.keys():
            count_dic_viewwise[views_count_std_key]['true'] = np.append(count_dic_viewwise[views_count_std_key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[views_count_std_key]['pred'] = np.append(count_dic_viewwise[views_count_std_key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[views_count_std_key]['prob'] = np.append(count_dic_viewwise[views_count_std_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[views_count_std_key]={}
            count_dic_viewwise[views_count_std_key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[views_count_std_key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[views_count_std_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    elif flag==1:
        views_count_4nonstd_key='views_nonstd'
        if views_count_4nonstd_key in count_dic_viewwise.keys():
            count_dic_viewwise[views_count_4nonstd_key]['true'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[views_count_4nonstd_key]['pred'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[views_count_4nonstd_key]['prob'] = np.append(count_dic_viewwise[views_count_4nonstd_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[views_count_4nonstd_key]={}
            count_dic_viewwise[views_count_4nonstd_key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[views_count_4nonstd_key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[views_count_4nonstd_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    if flag==0 and views_count_key=='4':
        views_count_4std_key='4_views_std'
        if views_count_4std_key in count_dic_viewwise.keys():
            count_dic_viewwise[views_count_4std_key]['true'] = np.append(count_dic_viewwise[views_count_4std_key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[views_count_4std_key]['pred'] = np.append(count_dic_viewwise[views_count_4std_key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[views_count_4std_key]['prob'] = np.append(count_dic_viewwise[views_count_4std_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[views_count_4std_key]={}
            count_dic_viewwise[views_count_4std_key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[views_count_4std_key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[views_count_4std_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    if views_names_key in count_dic_viewwise.keys():
        count_dic_viewwise[views_names_key]['true'] = np.append(count_dic_viewwise[views_names_key]['true'],test_labels.cpu().numpy())
        count_dic_viewwise[views_names_key]['pred'] = np.append(count_dic_viewwise[views_names_key]['pred'],test_pred.cpu().numpy())
        count_dic_viewwise[views_names_key]['prob'] = np.append(count_dic_viewwise[views_names_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
    else:
        count_dic_viewwise[views_names_key]={}
        count_dic_viewwise[views_names_key]['true'] = test_labels.cpu().numpy()
        count_dic_viewwise[views_names_key]['pred'] = test_pred.cpu().numpy()
        count_dic_viewwise[views_names_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    
    if views_count_key in count_dic_viewwise.keys():
        count_dic_viewwise[views_count_key]['true'] = np.append(count_dic_viewwise[views_count_key]['true'],test_labels.cpu().numpy())
        count_dic_viewwise[views_count_key]['pred'] = np.append(count_dic_viewwise[views_count_key]['pred'],test_pred.cpu().numpy())
        count_dic_viewwise[views_count_key]['prob'] = np.append(count_dic_viewwise[views_count_key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
    else:
        count_dic_viewwise[views_count_key]={}
        count_dic_viewwise[views_count_key]['true'] = test_labels.cpu().numpy()
        count_dic_viewwise[views_count_key]['pred'] = test_pred.cpu().numpy()
        count_dic_viewwise[views_count_key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    return count_dic_viewwise

def calc_viewwise_metric_newplot(views_names, views_standard, count_dic_viewwise, test_labels, test_pred, output_test):
    #view names concatenated together
    views_names_key="+".join(views_names)
    
    #number of views
    views_count_key=str(len(views_names))

    #number of views and number of breast sides
    breast_split = np.array([view[0] for view in views_names])
    breast_split = breast_split.tolist()

    flag=0
    for view in views_names:
        if view not in views_standard:
            flag=1
    
    #store all possible dictionary keys
    dic_keys = []
    if flag==0:
        dic_keys.append('views_std')
        if views_count_key == '4':
            dic_keys.append('4_views_std')
    else:
        dic_keys.append('views_nonstd')
    
    if (breast_split.count('L')>1) and (breast_split.count('R')==0):
        dic_keys.append('n>1L')
        dic_keys.append('oneside')
    if (breast_split.count('L')==0) and (breast_split.count('R')>1):
        dic_keys.append('n>1R')
        dic_keys.append('oneside')
    
    if views_count_key=='1':
        dic_keys.append('singleimage')
        if breast_split.count('L')==1:
            dic_keys.append('1L')
        elif breast_split.count('R')==1:
            dic_keys.append('1R')
    
    if (breast_split.count('L')==1) and (breast_split.count('R')==1):
        dic_keys.append('1L+1R')
    
    if (breast_split.count('L')>1) and (breast_split.count('R')>1):
        dic_keys.append('nL+mR')
    
    for key in dic_keys:
        if key in count_dic_viewwise.keys():
            count_dic_viewwise[key]['true'] = np.append(count_dic_viewwise[key]['true'],test_labels.cpu().numpy())
            count_dic_viewwise[key]['pred'] = np.append(count_dic_viewwise[key]['pred'],test_pred.cpu().numpy())
            count_dic_viewwise[key]['prob'] = np.append(count_dic_viewwise[key]['prob'],torch.sigmoid(output_test.data).cpu().numpy())
        else:
            count_dic_viewwise[key]={}
            count_dic_viewwise[key]['true'] = test_labels.cpu().numpy()
            count_dic_viewwise[key]['pred'] = test_pred.cpu().numpy()
            count_dic_viewwise[key]['prob'] = torch.sigmoid(output_test.data).cpu().numpy()
    
    return count_dic_viewwise

def write_results_viewwise(config_params, path_to_results_xlsx, sheetname, count_dic_viewwise):
    wb = op.load_workbook(path_to_results_xlsx)
    if count_dic_viewwise!={}:
        for key in count_dic_viewwise.keys():
            #if key=='4_views_std':
            print(key)
            per_model_metrics = aggregate_performance_metrics(config_params, count_dic_viewwise[key]['true'],count_dic_viewwise[key]['pred'],count_dic_viewwise[key]['prob'])
            header = [key]
            sheet = wb[sheetname]
            sheet.append(header)
            sheet.append(['Count','PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
            sheet.append([len(count_dic_viewwise[key]['true'])]+per_model_metrics)
    wb.save(path_to_results_xlsx)

def case_label_from_SIL(config_params, df_test, test_labels_all, test_pred_all, path_to_results):
    dic_true={}
    dic_pred={}
    idx=0
    
    if config_params['dataset'] == 'zgt':
        image_col = 'CasePath'
    elif config_params['dataset'] == 'cbis-ddsm':
        image_col = 'ImageName'
    elif config_params['dataset'] == 'vindr':
        image_col = 'StudyInstanceUID'
    elif config_params['dataset'] == 'cmmd':
        image_col = 'Patient_Id'

    test_pred_all = test_pred_all.reshape(-1)
    #print(test_labels_all)
    #print(test_pred_all)
    for idx in df_test.index:
        if config_params['dataset'] == 'cbis-ddsm':
            dic_key = '_'.join(df_test.loc[idx, image_col].split('_')[:3])
        elif config_params['dataset'] == 'zgt':
            dic_key = df_test.loc[idx, image_col].split('/')[-1]
        else:
            dic_key = df_test.loc[idx, image_col]
        dic_true[dic_key] = max(dic_true.get(dic_key,0), test_labels_all[idx])
        dic_pred[dic_key] = max(dic_pred.get(dic_key,0), test_pred_all[idx])
    case_labels_true = np.array(list(dic_true.values()))
    case_labels_pred = np.array(list(dic_pred.values()))

    #print(case_labels_true)
    #print(case_labels_pred)
    metrics_case_labels = aggregate_performance_metrics(config_params, case_labels_true, case_labels_pred, None)
    print("case label:", metrics_case_labels)
    write_results_xlsx(metrics_case_labels, path_to_results, 'test_results')

def data_specific_changes(config_params, df):
    if config_params['dataset'] == 'zgt':
        df = df.rename(columns = {'BreastDensity_standarized':'BreastDensity', 'BIRADS_combined_pathwaybased':'BIRADS'})
        df['BIRADS'] = df['BIRADS'].map({'1':'1', '2':'2', '3':'3', '4a':'4', '4b':'4', '4c':'4', '5':'5', '6':'6'})
    elif config_params['dataset'] == 'cbis-ddsm':
        df = df.rename(columns = {'AssessmentMax': 'BIRADS'})
        df['BreastDensity'] = df['BreastDensity'].map({1:'A', 2:'B', 3:'C', 4:'D'})
    return df

def results_breastdensity(config_params, df, true_labels, pred_labels, y_prob, path_to_results):
    df = data_specific_changes(config_params, df)
    
    breastden_A=df[df['BreastDensity']=='A'].index
    breastden_B=df[df['BreastDensity']=='B'].index
    breastden_C=df[df['BreastDensity']=='C'].index
    breastden_D=df[df['BreastDensity']=='D'].index
    
    breastdenA = aggregate_performance_metrics(config_params, true_labels[breastden_A], pred_labels[breastden_A], y_prob[breastden_A])
    breastdenB = aggregate_performance_metrics(config_params, true_labels[breastden_B], pred_labels[breastden_B], y_prob[breastden_B])
    breastdenC = aggregate_performance_metrics(config_params, true_labels[breastden_C], pred_labels[breastden_C], y_prob[breastden_C])
    breastdenD = aggregate_performance_metrics(config_params, true_labels[breastden_D], pred_labels[breastden_D], y_prob[breastden_D])

    results_all = [['Breast Density A'] + breastdenA, ['Breast Density B'] + breastdenB, ['Breast Density C'] + breastdenC, ['Breast Density D'] + breastdenD] 
    write_results_subgroup(path_to_results, 'BreastDensity', results_all)

def results_birads(config_params, df, true_labels, pred_labels, y_prob, path_to_results):
    df = data_specific_changes(config_params, df)

    if config_params['dataset'] == 'cbis-ddsm':
        birads_0=df[df['BIRADS']==0].index
        birads_1=df[df['BIRADS']==1].index
        birads_2=df[df['BIRADS']==2].index
        birads_3=df[df['BIRADS']==3].index
        birads_4=df[df['BIRADS']==4].index
        birads_5=df[df['BIRADS']==5].index
        birads_6=df[df['BIRADS']==6].index
    elif config_params['dataset'] == 'zgt':
        birads_0=df[df['BIRADS']=='0'].index
        birads_1=df[df['BIRADS']=='1'].index
        birads_2=df[df['BIRADS']=='2'].index
        birads_3=df[df['BIRADS']=='3'].index
        birads_4=df[df['BIRADS']=='4'].index
        birads_5=df[df['BIRADS']=='5'].index
        birads_6=df[df['BIRADS']=='6'].index
    
    print(birads_0.shape)
    print(birads_1.shape)
    print(birads_2.shape)
    print(birads_3.shape)
    print(birads_4.shape)
    print(birads_5.shape)
    print(birads_6.shape)

    try:
        birads0_res = aggregate_performance_metrics(config_params, true_labels[birads_0],pred_labels[birads_0],y_prob[birads_0])
    except:
        birads0_res = []
    try:
        birads1_res = aggregate_performance_metrics(config_params, true_labels[birads_1],pred_labels[birads_1],y_prob[birads_1])
    except:
        birads1_res = []
    try:
        birads2_res = aggregate_performance_metrics(config_params, true_labels[birads_2],pred_labels[birads_2],y_prob[birads_2])
    except:
        birads2_res = []
    try:
        birads3_res = aggregate_performance_metrics(config_params, true_labels[birads_3],pred_labels[birads_3],y_prob[birads_3])
    except:
        birads3_res = []
    try:
        birads4_res = aggregate_performance_metrics(config_params, true_labels[birads_4],pred_labels[birads_4],y_prob[birads_4])
    except:
        birads4_res = []
    try:
        birads5_res = aggregate_performance_metrics(config_params, true_labels[birads_5],pred_labels[birads_5],y_prob[birads_5])
    except:
        birads5_res = []
    try:
        birads6_res = aggregate_performance_metrics(config_params, true_labels[birads_6],pred_labels[birads_6],y_prob[birads_6])
    except:
        birads6_res = []
    
    results_all = [['BIRADS 0'] + birads0_res, ['BIRADS 1'] + birads1_res, ['BIRADS 2'] + birads2_res, ['BIRADS 3'] + birads3_res, ['BIRADS 4'] + birads4_res, ['BIRADS 5'] + birads5_res, ['BIRADS 6'] + birads6_res] 
    write_results_subgroup(path_to_results, 'BIRADS', results_all)

def results_abnormality(config_params, df, true_labels, pred_labels, y_prob, path_to_results):
    df = data_specific_changes(config_params, df)
    
    abnorm_mass = df[df['AbnormalityType']=='Mass'].index
    abnorm_calc = df[df['AbnormalityType']=='Calc'].index
    
    abnorm_mass_res = aggregate_performance_metrics(config_params, true_labels[abnorm_mass], pred_labels[abnorm_mass], y_prob[abnorm_mass])
    abnorm_calc_res = aggregate_performance_metrics(config_params, true_labels[abnorm_calc], pred_labels[abnorm_calc], y_prob[abnorm_calc])

    results_all = [['Mass'] + abnorm_mass_res, ['Calcification'] + abnorm_calc_res] 
    write_results_subgroup(path_to_results, 'Abnormality', results_all)

def write_results_subgroup(path_to_results, sheetname, results_all):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    sheet.append(['Group', 'PrecisionMicro','PrecisionMacro','RecallMicro','RecallMacro','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
    for result in results_all:
        sheet.append(result)
    wb.save(path_to_results)

def aggregate_performance_metrics(config_params, y_true, y_pred, y_prob): 
    try:
        prec_bin = metrics.precision_score(y_true, y_pred, average = 'binary')
    except:
        prec_bin = 0.0
    precmicro = metrics.precision_score(y_true, y_pred, average = 'micro')
    precmacro = metrics.precision_score(y_true, y_pred, average = 'macro')
    try:
        recall_bin = metrics.recall_score(y_true, y_pred, average = 'binary')
    except:
        recall_bin = 0.0
    recallmicro = metrics.recall_score(y_true, y_pred, average = 'micro')
    recallmacro = metrics.recall_score(y_true, y_pred, average = 'macro')
    try:
        f1_bin = metrics.f1_score(y_true, y_pred, average = 'binary')
    except:
        f1_bin = 0.0
    f1micro = metrics.f1_score(y_true, y_pred, average = 'micro')
    f1macro = metrics.f1_score(y_true, y_pred, average='macro')
    f1wtmacro=metrics.f1_score(y_true, y_pred, average='weighted')
    acc = metrics.accuracy_score(y_true, y_pred)
    cohen_kappa=metrics.cohen_kappa_score(y_true, y_pred)
    try:
        if len(config_params['classes']) > 2:
            auc = metrics.roc_auc_score(y_true, y_prob, average='macro', multi_class='ovo')
            auc_wtmacro = metrics.roc_auc_score(y_true, y_prob, average='weighted', multi_class='ovo')
        else:
            auc=metrics.roc_auc_score(y_true,y_prob)
            auc_wtmacro=0.0
    except:
        auc=0.0
        auc_wtmacro=0.0
    
    each_model_metrics=[prec_bin, precmicro, precmacro, recall_bin, recallmicro, recallmacro, f1_bin, f1micro, f1macro, f1wtmacro, acc, cohen_kappa, auc, auc_wtmacro]
    #don't uncomment this - I don't use the part below to log metrics to MLflow.
    #metrics_dic = {"Test precision positive class": prec_bin, "Test precision micro": precmicro, "Test precision macro": precmacro, "Test recall positive class": recall_bin, "Test recall micro": recallmicro, "Test recall macro": recallmacro, "Test f1 positive class": f1_bin, "Test f1 micro": f1micro, "Test f1 macro": f1macro, "Test f1 wt macro": f1wtmacro, "Test accuracy": acc, "Test cohen kappa": cohen_kappa, "Test AUC": auc, "Test AUC wt macro": auc_wtmacro}
    #with mlflow.start_run():
    #mlflow.log_metrics(metrics_dic)
    return each_model_metrics

def classspecific_performance_metrics(config_params, y_true, y_pred, y_prob, path_to_results, sheetname):
    score_dict = classification_report(y_true, y_pred, labels=config_params['classes'], output_dict = True)
    print(score_dict)
    results_all = []
    flag=0
    for key in score_dict.keys():
        if isinstance(score_dict[key], dict):
            if flag == 0:
                results_all.append(['class'] + list(score_dict[key].keys()))
                flag = 1
            results_all.append([key] + list(score_dict[key].values())) 
        else:
            results_all.append([key, score_dict[key]])
    
    print(results_all)
    write_results_classspecific(path_to_results, sheetname, results_all)

def write_results_classspecific(path_to_results, sheetname, results_all):
    wb = op.load_workbook(path_to_results)
    if sheetname not in wb.sheetnames:
        sheet = wb.create_sheet(sheetname)
    else:
        sheet = wb[sheetname]
    for result in results_all:
        sheet.append(result)
    wb.save(path_to_results)
  