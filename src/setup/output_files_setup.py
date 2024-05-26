# -*- coding: utf-8 -*-
"""
Created on Wed Sep 22 16:43:24 2021

@author: PathakS
"""

import os
import sys
import openpyxl as op
from openpyxl import Workbook

def output_files(config_file, config_params, num_config_start, num_config_end):
    #output files path
    path_to_output = "/".join(config_file.split('/')[:-1])+'/'

    #check output_folder path
    if not os.path.exists(path_to_output):
        print("Error! config file path does not exist! This code needs the same path to store the output files and model.")
        sys.exit()
    
    if config_params['randseeddata']!=config_params['randseedother']:
        rand_seed = str(config_params['randseedother']) +'_'+ str(config_params['randseeddata'])
    else:
        rand_seed = str(config_params['randseeddata'])
    
    path_to_hyperparam_search = path_to_output+"hyperparamsearch_"+str(num_config_start)+'-'+str(num_config_end)+'_'+str(rand_seed)+".xlsx"
    
    if config_params['run']:
        path_to_model = path_to_output+"model_"+str(rand_seed)+'_'+config_params['run']+".tar"
        path_to_results = path_to_output+"result_"+str(rand_seed)+'_'+config_params['run']+".xlsx"
        path_to_results_text = path_to_output+"result_"+str(rand_seed)+'_'+config_params['run']+".txt"
        path_to_learning_curve = path_to_output+"learningcurve_"+str(rand_seed)+'_'+config_params['run']+".png"
        path_to_log_file = path_to_output+"log_"+str(rand_seed)+'_'+config_params['run']+".txt"

    else:
        #path_to_model = "/homes/spathak/multiview_mammogram/models_results/vindr/ijcai23/modelid15_viewsinclusionall_femodelgmic_resnet18_learningtypeSIL/model_8.tar"
        #path_to_model = "/homes/spathak/multiview_mammogram/models_results/cbis-ddsm/paper2/modelid20_viewsinclusionall_femodelgmic_resnet18_learningtypeSIL/model_8_1.tar"
        path_to_model = path_to_output+"model_"+str(rand_seed)+".tar"
        path_to_results = path_to_output+"result_"+str(rand_seed)+".xlsx" #"_trainVinDrtestZGT.xlsx" #_testMGMFV.xlsx
        path_to_results_text = path_to_output+"result_"+str(rand_seed)+".txt" #"_trainVinDrtestZGT.txt"
        path_to_learning_curve = path_to_output+"learningcurve_"+str(rand_seed)+".png" #"_trainVinDrtestZGT.png"
        path_to_log_file = path_to_output+"log_"+str(rand_seed)+".txt" #"_trainVinDrtestZGT.txt"
    
    # set file path
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        sheet1 = wb['train_val_results']
        sheet2 = wb['confmat_train_val_test']
        sheet3 = wb['test_results']
        sheet4 = wb['metrics_view_wise']
    else:
        wb=Workbook()
        sheet1 = wb.active
        sheet1.title="train_val_results"
        if config_params['usevalidation']:
            header=['Epoch','lr','Avg Loss Train','Accuracy Train','F1macro Train','Recall Train','Speci Train','Avg Loss Val','Accuracy Val','F1macro Val','Recall Val','Speci Val','AUC Val', 'AUC WtMacro Val']
        else:
            header=['Epoch','lr','Avg Loss Train','Accuracy Train','F1macro Train','Recall Train','Speci Train']+['Loss','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC']
        sheet1.append(header)
        sheet2 = wb.create_sheet('confmat_train_val_test')
        sheet3 = wb.create_sheet('test_results') 
        #sheet3.append(['Loss','Precision','Recall','Specificity','F1','F1macro','F1wtmacro','Acc','Bal_Acc','Cohens Kappa','AUC'])
        sheet3.append(['Epoch','Loss','PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC','AUCWtMacro'])
        sheet4 = wb.create_sheet('metrics_view_wise')
    
    # set file path
    if config_params['usevalidation']:
        if os.path.isfile(path_to_hyperparam_search):
            wb1 = op.load_workbook(path_to_hyperparam_search)
            sheet5 = wb1['hyperparam_results']
        else:
            wb1=Workbook()
            sheet5 = wb1.active
            sheet5.title = "hyperparam_results"
            header = ['config_file','lr','wtdecay','sm_reg_param','trainingscheme','optimizer','patienceepoch','batchsize','Loss','PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC']
            sheet5.append(header)
        wb1.save(path_to_hyperparam_search)

    wb.save(path_to_results)

    if config_params['usevalidation']:
        return path_to_model, path_to_results, path_to_results_text, path_to_learning_curve, path_to_log_file, path_to_hyperparam_search 
    else:
        return path_to_model, path_to_results, path_to_results_text, path_to_learning_curve, path_to_log_file


    
        
        
        
        
        
        
        