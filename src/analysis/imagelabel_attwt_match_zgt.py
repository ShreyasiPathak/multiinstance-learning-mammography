import torch
import numpy as np
import pandas as pd
import openpyxl as op
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix

from utilities import data_loaders_utils
from train_eval import test, evaluation

pd.options.mode.chained_assignment = None

def imglabel_f1_plot(filename):
    fig, ax = plt.subplots(figsize=(23,15))
    
    x = np.arange(8)
    width = 0.25  # the width of the bars
    multiplier = 0
    
    labels = ['SIL-CL', 'IS-Mean$^{img}$', 'IS-Att$^{img}$', 'IS-GAtt$^{img}$', 'IS-Att$^{side}$', 'ES-Att$^{img}$', 'ES-GAtt$^{img}$', 'ES-Att$^{side}$']
    
    #meanF1 = [0.70, 0.62, 0.52, 0.64, 0.73, 0.70, 0.76, 0.50, 0.81, 0.76, 0.74, 0.81]
    #stddev = [0.03, 0.01, 0.0, 0.02, 0.01, 0.01, 0.02, 0.02, 0.02, 0.0, 0.01, 0.04]

    meanF1_sil = (0.64)
    #std_sil = (0.03, 0.02)
    x_sil = np.arange(1)
    #label_es = ['AttWt', 'AttWt', 'AttWt']
    rects = ax.bar(x_sil + width, meanF1_sil, width, color = 'dodgerblue', capsize = 2)
    ax.bar_label(rects, padding=3, fontsize=15)

    meanF1_is = {'Image Probability': (0.66, 0.78, 0.77, 0.72), 'Attention Score': (0.43, 0.69, 0.66, 0.76)}
    x_is = np.arange(1,5)
    colour_select = ['dodgerblue', 'orange']
    for attribute, measurement in meanF1_is.items():
        offset = width * multiplier
        rects = ax.bar(x_is + offset, measurement, width, label=attribute, color = colour_select[multiplier], capsize = 2)
        ax.bar_label(rects, padding=3, fontsize=15)
        multiplier += 1
    
    meanF1_es = (0.75, 0.74, 0.76)
    x_es = np.arange(5,8)
    rects = ax.bar(x_es+width, meanF1_es, width, color ='orange', capsize = 2)
    ax.bar_label(rects, padding=3, fontsize=15)

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('F1$^{\mu}$ score at the image-level', fontsize=25)
    ax.set_xlabel('Model variants', fontsize=25)
    #ax.set_title('Penguin attributes by species')
    ax.tick_params(axis='y', which='major', labelsize=25)
    ax.set_xticks(x + width, labels, fontsize = 25)
    ax.legend(loc='upper left', fontsize=25)
    fig.savefig(filename)

def match_imagelevel_groundtruth_withattnwt(config_params, exam_name, attwt, instance_prob, img_path_list, df_image):
    #print(exam_name)
    #zgt_roi_annotation = '/homes/spathak/multiview_mammogram/input_data/extracted_annotations_withrle_corrected.csv'
    #df_grp = df_image[df_image['StudyInstanceUID']==exam_name].sort_values(by='Views')
    #if config_params['learningtype'] == 'MIL':
    #    df_grp = df_image[df_image['image'].str.split('/').str[-3]==exam_name]
    #elif config_params['learningtype'] == 'SIL':
    df_grp = df_image[df_image['ShortPath'].str.split('/').str[-3]==exam_name]
    #if np.unique(df_grp['CaseLabel'])[0] == 'malignant':
    print(exam_name, flush=True) #case name: studyinstanceuid_accessionnum
    #print(df_grp, flush=True)
    print(df_grp['diagnosis'], flush=True)
    df_grp['TrueImageLabel'] = df_grp['diagnosis'].map({'benign':0, 'malignant':1})
    print(df_grp['TrueImageLabel'])
    #print("df grp image:", df_grp['ShortPath'].values, flush=True)
    #print("img path list:", img_path_list, flush=True)
    df_grp['modifiedname'] = config_params['preprocessed_imagepath']+'/' +df_grp['ShortPath'].str.split('.png', regex=False).str[0] + '.npy'
    #print(df_grp['modifiedname'].values.tolist())
    
    assert df_grp['modifiedname'].values.tolist() == img_path_list
    #assert attwt[0].shape[0] == 4
    try:
        impscore = 1/attwt[0].shape[0]
    except:
        impscore = 1/df_grp['TrueImageLabel'].shape[0]
    print("importance score:", impscore)
    
    if config_params['milpooling'] == 'ismean':
        attwt =  np.array([[impscore] * df_grp['TrueImageLabel'].shape[0]])
        print(attwt, flush=True)
        #attwt = np.array([[impscore, impscore, impscore, impscore]])

    if config_params['milpooling'] == 'ismean' or config_params['milpooling'] == 'isatt' or config_params['milpooling'] == 'isgatt':
        if config_params['attention'] == 'breastwise':
            #print("instance prob breastwise:", instance_prob, flush=True)
            #print(instance_prob[0].shape)
            #print(instance_prob[1].shape)
            instance_prob = torch.cat((instance_prob[0].view(1,-1), instance_prob[1].view(1,-1)), dim=1)
            #print("instance prob breastwise:", instance_prob, flush=True)

        instance_prob = instance_prob.view(1,-1).cpu().numpy()
        print("instance prob:", instance_prob, flush=True)
        df_grp['InstanceProb'] = (np.array(list(instance_prob[0]))>0.50).astype(int)
        print(df_grp['InstanceProb'], flush=True)
        
        print("att wt:", attwt, flush=True)
        df_grp['ModelAttwt'] = (np.array(list(attwt[0]))>impscore).astype(int) #>=impscore for IS-Mean model of ZGT data
        print(df_grp['ModelAttwt'], flush=True)

        image_wt = np.multiply(instance_prob, attwt)
        print("image wt:", image_wt, flush=True)
        df_grp['InstanceImportance'] = (np.array(list(image_wt[0]))>impscore).astype(int)
        print(df_grp['InstanceImportance'], flush=True)
        
    elif config_params['milpooling'] == 'esmean' or config_params['milpooling'] == 'esatt' or config_params['milpooling'] == 'esgatt':
        print("att wt:", attwt, flush=True)
        df_grp['ModelAttwt'] = (np.array(list(attwt[0]))>impscore).astype(int)
    
    #print(df_grp[['StudyInstanceUID','Views','ImageLabel', 'CaseLabel','ModelAttwt']])
    #y_true = df_grp['TrueImageLabel'].values.tolist()
    #y_pred = df_grp['ModelAttwt'].values.tolist()
    #return y_true, y_pred
    return df_grp
    
def extract_img_attn_wts(config_params, img_attns):
    if config_params['attention'] == 'imagewise':
        if config_params['milpooling'] == 'ismean':
            A_all = None
        else:
            A_all = img_attns.view(img_attns.shape[0], -1).data.cpu().numpy()
    elif config_params['attention'] == 'breastwise':
        A_left = img_attns[0]
        A_right = img_attns[1]
        A_both = img_attns[2]
        '''print(A_left)
        print(A_left.shape)
        print(A_right)
        print(A_right.shape)
        print(A_both)
        print(A_both.shape)
        '''

        if (A_left is not None) and (A_right is not None) and (A_both is not None):
            A_left_both = torch.mul(A_left, A_both[:,:,0])
            A_right_both = torch.mul(A_right, A_both[:,:,1])
            A_all = torch.cat((A_left_both, A_right_both), dim=2)
            A_all = A_all.view(1, -1).data.cpu().numpy()
        elif (A_left is None) and (A_right is None) and (A_both is not None):
            A_all = A_both.view(1, -1).data.cpu().numpy()
        elif (A_left is not None) and (A_right is None) and (A_both is None):
            A_all = A_left.view(1, -1).data.cpu().numpy()
        elif (A_left is None) and (A_right is not None) and (A_both is None):
            A_all = A_right.view(1, -1).data.cpu().numpy()

        #print(A_all)
        #input('halt')
    return A_all

def group_to_row(df_grp1):
    df_grp2 = pd.DataFrame(columns=['image', 'ShortPath', 'diagnosis'], index=range(1))
    if 'Malignant' in df_grp1['diagnosis'].values.tolist():
        df_grp2['diagnosis'] = 'malignant'
    else:
        df_grp2['diagnosis'] = 'benign'
    
    df_grp2['image'] = df_grp1['image'].unique().item()
    df_grp2['ShortPath'] = df_grp1['ShortPath'].unique().item()
    return df_grp2

def model_output(config_params, model, dataloader_test, df_test, path_to_results):
    model.eval()
    eval_mode = True
    count = 0
    pval_sum =0
    corr_sum =0
    y_true_all = []
    y_pred_insprob_all = []
    y_pred_attwt_all = []
    y_pred_insimp_all = []
    y_pred_all = []

    #zgt_roi_annotation = '/homes/spathak/multiview_mammogram/input_data/extracted_annotations_withrle_corrected.csv'
    zgt_roi_annotation = '/home/pathaks/PhD/case-level-breast-cancer/multiview_mammogram/input_data/extracted_annotations_withrle_corrected.csv'
    df_roi = pd.read_csv(zgt_roi_annotation, sep=';')
    df_roi = df_roi[['image','diagnosis']]
    df_roi['ShortPath'] = '/' + df_roi['image'].str.split('=').str[-1]
    #print(df_roi.shape, flush=True)
    df_roi = df_roi.groupby(by='ShortPath', as_index=False).apply(group_to_row)
    #print(df_roi.shape, flush=True)
    #print(df_roi)
    #print(df_roi['ShortPath'], flush=True)
    if config_params['learningtype'] == 'MIL':
        df_si = pd.read_csv(config_params['SIL_csvfilepath'], sep=';')
        df_image = df_si.merge(df_roi, on=['ShortPath'], how='inner').sort_values(by=['Patient_Id', 'AccessionNum', 'Views'])
        #print(df_image.columns)
    elif config_params['learningtype'] == 'SIL':
        print(df_test, flush=True)
        df_test = df_test.merge(df_roi, on=['ShortPath'], how='left')#.sort_values(by=['Patient_Id', 'AccessionNum', 'Views'])
        print(df_test, flush=True)
        

    with torch.no_grad():
        for test_idx, test_batch, test_labels, views_names in dataloader_test:
            test_batch, test_labels = test_batch.to(config_params['device']), test_labels.to(config_params['device'])
            test_labels = test_labels.view(-1)
            print("test idx:", test_idx.item(), flush=True)
            print("test batch:", test_batch.shape, flush=True)
            if config_params['femodel'] == 'gmic_resnet18':
                if config_params['learningtype'] == 'SIL':
                    loaded_image = data_loaders_utils.collect_images(config_params, df_test.loc[test_idx.item()])
                    _, _, output_batch_fusion, saliency_map, patch_locations, patch_imgs, patch_attns, _ = model(test_batch, eval_mode) # compute model output, loss and total train loss over one epoch
                    loaded_image = loaded_image[np.newaxis,:,:,:]
                    patch_locations = patch_locations[:,np.newaxis,:]
                    patch_imgs = patch_imgs[:,np.newaxis,:]
                    patch_attns = patch_attns[:,np.newaxis,:]
                    saliency_map = saliency_map[:, np.newaxis, :, :, :]
                    img_attns = None
                    exam_name = df_test.loc[test_idx.item(), 'ShortPath'].split('/')[-1]
                    #print(exam_name)

                elif config_params['learningtype'] == 'MIL':
                    if config_params['dataset'] == 'cbis-ddsm':
                        exam_name = df_test.loc[test_idx.item(), 'FolderName']
                    elif config_params['dataset'] == 'vindr':
                        exam_name = df_test.loc[test_idx.item(), 'StudyInstanceUID']
                    elif config_params['dataset'] == 'zgt':
                        exam_name = df_test.loc[test_idx.item(), 'ShortPath'].split('/')[-1]
                    loaded_image, _, _, img_path_list = data_loaders_utils.collect_cases(config_params, df_test.loc[test_idx.item()])
                    _, _, output_batch_fusion, saliency_map, patch_locations, patch_imgs, patch_attns, img_attns, _, instance_prob = model(test_batch, views_names, eval_mode)
            
            test_pred = torch.ge(torch.sigmoid(output_batch_fusion), torch.tensor(0.5)).float()
            #print(list(test_labels.cpu().numpy()))
            #print(list(test_pred.cpu().numpy()[0].astype(int)))
            saliency_maps = saliency_map.cpu().numpy()
            patch_attentions = patch_attns[0, :, :].data.cpu().numpy()

            if config_params['learningtype'] == 'SIL':
                if df_test.loc[test_idx.item(), 'CaseLabel']=='malignant':
                    if config_params['labeltouse'] == 'imagelabel':
                        y_true_all.append(list(test_labels.cpu().numpy()))
                    elif config_params['labeltouse'] == 'caselabel':
                        y_true_all.append(config_params['groundtruthdic'][df_test.loc[test_idx.item(), 'diagnosis']])
                    y_pred_all.append(list(test_pred.cpu().numpy()[0].astype(int)))
                    count+=1

            elif config_params['learningtype'] == 'MIL':
                if test_labels.item() == 1:
                    img_attns = extract_img_attn_wts(config_params, img_attns)
                    df_grp = match_imagelevel_groundtruth_withattnwt(config_params, exam_name, img_attns, instance_prob, img_path_list, df_image)
                    if config_params['milpooling'] == 'ismean' or config_params['milpooling'] == 'isatt' or config_params['milpooling'] == 'isgatt':
                        y_true = df_grp['TrueImageLabel'].values.tolist()
                        y_pred_insprob = df_grp['InstanceProb'].values.tolist()
                        y_pred_attwt = df_grp['ModelAttwt'].values.tolist()
                        y_pred_insimp = df_grp['InstanceImportance'].values.tolist()
                        y_true_all.extend(y_true)
                        y_pred_insprob_all.extend(y_pred_insprob)
                        y_pred_attwt_all.extend(y_pred_attwt)
                        y_pred_insimp_all.extend(y_pred_insimp)
                    
                    elif config_params['milpooling'] == 'esmean' or config_params['milpooling'] == 'esatt' or config_params['milpooling'] == 'esgatt':
                        y_true = df_grp['TrueImageLabel'].values.tolist()
                        y_pred = df_grp['ModelAttwt'].values.tolist()
                        y_true_all.extend(y_true)
                        y_pred_all.extend(y_pred)
                count+=1
    
    print("count:", count, flush=True)
    
    if config_params['learningtype'] == 'MIL':
        if config_params['milpooling'] == 'ismean' or config_params['milpooling'] == 'isatt' or config_params['milpooling'] == 'isgatt':
            y_true_all = np.array(y_true_all)
            
            y_pred_insprob_all = np.array(y_pred_insprob_all)
            conf_mat_insprob = confusion_matrix(y_true_all, y_pred_insprob_all)
            print(conf_mat_insprob)
            evaluation.write_results_xlsx_confmat(config_params, conf_mat_insprob, path_to_results, 'imglabel_confmat')
            wb = op.load_workbook(path_to_results)
            sheet = wb['imglabel_confmat']
            sheet.append(['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
            wb.save(path_to_results)
            per_model_metric = evaluation.aggregate_performance_metrics(config_params, y_true_all, y_pred_insprob_all, None)
            print(per_model_metric)
            evaluation.write_results_xlsx(per_model_metric, path_to_results, 'imglabel_confmat')

            y_pred_attwt_all = np.array(y_pred_attwt_all)
            conf_mat_attwt = confusion_matrix(y_true_all, y_pred_attwt_all)
            print(conf_mat_attwt)
            evaluation.write_results_xlsx_confmat(config_params, conf_mat_attwt, path_to_results, 'imglabel_confmat')
            wb = op.load_workbook(path_to_results)
            sheet = wb['imglabel_confmat']
            sheet.append(['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
            wb.save(path_to_results)
            per_model_metric = evaluation.aggregate_performance_metrics(config_params, y_true_all, y_pred_attwt_all, None)
            print(per_model_metric)
            evaluation.write_results_xlsx(per_model_metric, path_to_results, 'imglabel_confmat')

            y_pred_insimp_all = np.array(y_pred_insimp_all)
            conf_mat_insimp = confusion_matrix(y_true_all, y_pred_insimp_all)
            print(conf_mat_insimp)
            evaluation.write_results_xlsx_confmat(config_params, conf_mat_insimp, path_to_results, 'imglabel_confmat')
            wb = op.load_workbook(path_to_results)
            sheet = wb['imglabel_confmat']
            sheet.append(['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
            wb.save(path_to_results)
            per_model_metric = evaluation.aggregate_performance_metrics(config_params, y_true_all, y_pred_insimp_all, None)
            print(per_model_metric)
            evaluation.write_results_xlsx(per_model_metric, path_to_results, 'imglabel_confmat')
        
        elif config_params['milpooling'] == 'esmean' or config_params['milpooling'] == 'esatt' or config_params['milpooling'] == 'esgatt':
            y_true_all = np.array(y_true_all)
            y_pred_all = np.array(y_pred_all)
            conf_mat = confusion_matrix(y_true_all, y_pred_all)
            print(conf_mat)
            evaluation.write_results_xlsx_confmat(config_params, conf_mat, path_to_results, 'imglabel_confmat')
            print(count)
            wb = op.load_workbook(path_to_results)
            sheet = wb['imglabel_confmat']
            sheet.append(['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
            wb.save(path_to_results)
            per_model_metric = evaluation.aggregate_performance_metrics(config_params, y_true_all, y_pred_all, None)
            print(per_model_metric)
            evaluation.write_results_xlsx(per_model_metric, path_to_results, 'imglabel_confmat')
    elif config_params['learningtype'] == 'SIL':
        y_true_all = np.array(y_true_all)
        y_pred_all = np.array(y_pred_all)
        conf_mat = confusion_matrix(y_true_all, y_pred_all)
        print(conf_mat)
        evaluation.write_results_xlsx_confmat(config_params, conf_mat, path_to_results, 'imglabel_confmat')
        print(count)
        wb = op.load_workbook(path_to_results)
        sheet = wb['imglabel_confmat']
        sheet.append(['PrecisionBin','PrecisionMicro','PrecisionMacro','RecallBin','RecallMicro','RecallMacro','F1Bin','F1Micro','F1macro','F1wtmacro','Acc','Cohens Kappa','AUC'])
        wb.save(path_to_results)
        per_model_metric = evaluation.aggregate_performance_metrics(config_params, y_true_all, y_pred_all, None)
        print(per_model_metric)
        evaluation.write_results_xlsx(per_model_metric, path_to_results, 'imglabel_confmat')

def run_imagelabel_attwt_match(config_params, model, path_to_model, dataloader_test, df_test, path_to_results):
    path_to_trained_model = path_to_model
    model1 = test.load_model_for_testing(model, path_to_trained_model)
    model_output(config_params, model1, dataloader_test, df_test, path_to_results)

#filename = 'C:/Users/PathakS/OneDrive - Universiteit Twente/PhD/projects/radiology breast cancer/codes/breast-cancer-multiview-mammogram-codes/multiinstance results/results/NextSubmission/f1_attmodel_gt_imglabel_zgt1.pdf'
#meanF1 = [0.52, 0.75, 0.75, 0.75, 0.72, 0.82]
#stddev = [0, 0.03, 0.02, 0.01, 0, 0.04]
#meanF1 = [0.70, 0.62, 0.52, 0.64, 0.73, 0.70, 0.76, 0.50, 0.81, 0.76, 0.74, 0.81]
#stddev = [0.03, 0.01, 0.0, 0.02, 0.01, 0.01, 0.02, 0.02, 0.02, 0.0, 0.01, 0.04]
#imglabel_f1_plot(filename)