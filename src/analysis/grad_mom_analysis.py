import os
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as op

def plot_grad_dic_flow(case_dic, figurename, config_params):
    #plt.plot([case_dic[1]/case_size.count(1), case_dic[2]/case_size.count(2), case_dic[3]/case_size.count(3), case_dic[4]/case_size.count(4), case_dic[5]/case_size.count(5), case_dic[6]/case_size.count(6), case_dic[7]/case_size.count(0)], alpha=0.3, color="b")
    plt.clf()
    plt.plot(list(case_dic.values()), 'o-b')
    #plt.hlines(0, 0, len(list(case_dic.keys()))+1, linewidth=1, color="k" )
    plt.xticks(range(0,len(list(case_dic.keys())), 1), list(case_dic.keys()), rotation=25, ha='center')
    plt.xlim(xmin=0, xmax=len(list(case_dic.keys())))
    plt.xlabel("Case groups")
    plt.ylabel("Average gradient")
    plt.title("Gradient across case groups")
    plt.tight_layout()
    plt.savefig(os.path.join(config_params['path_to_output'], figurename), format='pdf')
    #plt.grid(True)

def store_grad_flow(named_parameters, layer_name):
    layers = []
    sum_grad = 0.0
    for name, param in named_parameters:
        if (param.requires_grad) and (layer_name in name):
            layers.append(name)
            #sum_grad+=param.grad.abs().mean()
            sum_grad+=param.grad.mean()
    #case_size.append(len(view_names))
    #ave_grads.append(sum_grad/len(layers))
    ave_grads = sum_grad/len(layers)
    ave_grads = ave_grads.cpu().item()
    return ave_grads

def grad_analysis(model, config_params, views_names, case_size, case_grad_dic_img, case_grad_dic_side, epoch, batch_no):
    #number of views and number of breast sides
    breast_split = np.array([view[0] for view in views_names])
    breast_split = breast_split.tolist()
    dic_key = 'L-'+str(breast_split.count('L'))+'+'+'R-'+str(breast_split.count('R'))
    case_size.append(dic_key)
    
    #img.attention plot
    avg_grad_img = store_grad_flow(model.named_parameters(), 'img.attention')
    if dic_key in case_grad_dic_img.keys():
        case_grad_dic_img[dic_key] = case_grad_dic_img[dic_key] + avg_grad_img
    else:
        case_grad_dic_img[dic_key] = 0
        case_grad_dic_img[dic_key] = case_grad_dic_img[dic_key] + avg_grad_img
    
    case_dic_img = dict()
    #print("before:", case_grad_dic_img)
    for key in case_grad_dic_img.keys():
        case_dic_img[key] = case_grad_dic_img[key]/case_size.count(key)
    #print("after:", case_dic_img)
    plot_grad_dic_flow(case_dic_img, 'casegroup_gradient_img_epoch'+str(epoch)+'.pdf', config_params)

    #side.attention plot
    avg_grad_side = store_grad_flow(model.named_parameters(), 'side.attention')
    if dic_key in case_grad_dic_side.keys():
        case_grad_dic_side[dic_key] = case_grad_dic_side[dic_key] + avg_grad_side
    else:
        case_grad_dic_side[dic_key] = 0
        case_grad_dic_side[dic_key] = case_grad_dic_side[dic_key] + avg_grad_side
    
    case_dic_side = dict()
    #print("before:", case_grad_dic_side)
    for key in case_grad_dic_side.keys():
        case_dic_side[key] = case_grad_dic_side[key]/case_size.count(key)
    #print("after:", case_dic_side)
    plot_grad_dic_flow(case_dic_side, 'casegroup_gradient_side_epoch'+str(epoch)+'.pdf', config_params)

    wb_grad = op.load_workbook(os.path.join(config_params['path_to_output'], 'average_grad_epoch'+str(epoch)+'.xlsx'))
    sheet_grad = wb_grad['results']
    sheet_grad.append([batch_no, '+'.join(views_names), avg_grad_img, avg_grad_side])
    wb_grad.save((os.path.join(config_params['path_to_output'], 'average_grad_epoch'+str(epoch)+'.xlsx')))

    return case_size, case_grad_dic_img, case_grad_dic_side

def plot_momentum_flow(momentum, figurename, config_params):
    #plt.plot([case_dic[1]/case_size.count(1), case_dic[2]/case_size.count(2), case_dic[3]/case_size.count(3), case_dic[4]/case_size.count(4), case_dic[5]/case_size.count(5), case_dic[6]/case_size.count(6), case_dic[7]/case_size.count(0)], alpha=0.3, color="b")
    plt.clf()
    plt.plot(momentum)
    #plt.hlines(0, 0, len(list(case_dic.keys()))+1, linewidth=1, color="k" )
    #plt.xticks(range(0,len(list(case_dic1.keys())), 1), list(case_dic1.keys()), rotation=25, ha='center')
    #plt.xlim(xmin=0, xmax=len(list(case_dic1.keys())))
    plt.xlabel("Batch iteration")
    plt.ylabel("Average momentum")
    plt.title("Momentum update")
    plt.tight_layout()
    plt.savefig(os.path.join(config_params['path_to_output'], figurename))
    #plt.grid(True)

def plot_momentum_flow_group(case_dic, figurename, config_params):
    #plt.plot([case_dic[1]/case_size.count(1), case_dic[2]/case_size.count(2), case_dic[3]/case_size.count(3), case_dic[4]/case_size.count(4), case_dic[5]/case_size.count(5), case_dic[6]/case_size.count(6), case_dic[7]/case_size.count(0)], alpha=0.3, color="b")
    plt.clf()
    plt.plot(list(case_dic.values()), 'o-b')
    #plt.hlines(0, 0, len(list(case_dic.keys()))+1, linewidth=1, color="k" )
    plt.xticks(range(0,len(list(case_dic.keys())), 1), list(case_dic.keys()), rotation=25, ha='center')
    plt.xlim(xmin=0, xmax=len(list(case_dic.keys())))
    plt.xlabel("Case groups")
    plt.ylabel("Average momentum")
    plt.title("Momentum across case groups")
    plt.tight_layout()
    plt.savefig(os.path.join(config_params['path_to_output'], figurename))
    #plt.grid(True)

def momentum_analysis(optimizer, config_params, views_names, case_mom_size, case_mom_dic_img, case_mom_dic_side, avg_mom_img, avg_mom_side, epoch, batch_no):
    #print("param_groups:", len(optimizer.param_groups))
    #print("state dict:", optimizer.state_dict())
    if config_params['attention'] == 'breastwise' and (config_params['milpooling'] == 'esatt' or config_params['milpooling'] == 'esgatt' or config_params['milpooling'] == 'isatt' or config_params['milpooling'] == 'isgatt'):
        optimizer_params_dic = {'img.attention':0, 'side.attention':1}
    elif config_params['attention'] == 'imagewise' and (config_params['milpooling'] == 'esatt' or config_params['milpooling'] == 'esgatt' or config_params['milpooling'] == 'isatt' or config_params['milpooling'] == 'isgatt'):
        optimizer_params_dic = {'img.attention':0}

    #number of views and number of breast sides
    breast_split = np.array([view[0] for view in views_names])
    breast_split = breast_split.tolist()
    dic_key = 'L-'+str(breast_split.count('L'))+'+'+'R-'+str(breast_split.count('R'))
    case_mom_size.append(dic_key)

    #momentum plot img.attention block
    sum_mom_img=0
    value = optimizer_params_dic['img.attention']
    for id in optimizer.state_dict()['param_groups'][value]['params']:
        sum_mom_img+=optimizer.state_dict()['state'][id]['exp_avg'].mean()
    avg_mom_block_img = (sum_mom_img/len(optimizer.state_dict()['param_groups'][value]['params'])).cpu().item()
    avg_mom_img.append(avg_mom_block_img)
    plot_momentum_flow(avg_mom_img, 'momentum_img.pdf', config_params)
    
    if dic_key in case_mom_dic_img.keys():
        case_mom_dic_img[dic_key] = case_mom_dic_img[dic_key] + avg_mom_block_img
    else:
        case_mom_dic_img[dic_key] = 0
        case_mom_dic_img[dic_key] = case_mom_dic_img[dic_key] + avg_mom_block_img
    
    case_dic_img = dict()
    for key in case_mom_dic_img.keys():
        case_dic_img[key] = case_mom_dic_img[key]/case_mom_size.count(key)
    plot_momentum_flow_group(case_dic_img, 'casegroup_momentum_img_epoch'+str(epoch)+'.pdf', config_params)

    #momentum plot side.attention block
    sum_mom_side=0
    value = optimizer_params_dic['side.attention']
    for id in optimizer.state_dict()['param_groups'][value]['params']:
        sum_mom_side+=optimizer.state_dict()['state'][id]['exp_avg'].mean()
    avg_mom_block_side = (sum_mom_side/len(optimizer.state_dict()['param_groups'][value]['params'])).cpu().item()
    avg_mom_side.append(avg_mom_block_side)
    plot_momentum_flow(avg_mom_side, 'momentum_side.pdf', config_params)

    if dic_key in case_mom_dic_side.keys():
        case_mom_dic_side[dic_key] = case_mom_dic_side[dic_key] + avg_mom_block_side
    else:
        case_mom_dic_side[dic_key] = 0
        case_mom_dic_side[dic_key] = case_mom_dic_side[dic_key] + avg_mom_block_side
    
    case_dic_side = dict()
    for key in case_mom_dic_side.keys():
        case_dic_side[key] = case_mom_dic_side[key]/case_mom_size.count(key)
    
    plot_momentum_flow_group(case_dic_side, 'casegroup_momentum_side_epoch'+str(epoch)+'.pdf', config_params)
    
    #mom_dic = {'BatchNum':batch_no, 'ViewName': views_names, 'AvgMomImg': avg_mom_block_img.cpu(), 'AvgMomSide': avg_mom_block_side.cpu()}
    #writer.writerow(mom_dic)
    
    wb_mom = op.load_workbook(os.path.join(config_params['path_to_output'], 'average_mom_epoch'+str(epoch)+'.xlsx'))
    sheet_mom = wb_mom['results']
    sheet_mom.append([batch_no, '+'.join(views_names), avg_mom_block_img, avg_mom_block_side])
    wb_mom.save((os.path.join(config_params['path_to_output'], 'average_mom_epoch'+str(epoch)+'.xlsx')))

    return case_mom_size, case_mom_dic_img, case_mom_dic_side, avg_mom_img, avg_mom_side

def momentum_analysis_fixedview(model, optimizer, config_params, views_names, case_mom_size, case_mom_dic_img, case_mom_dic_side, avg_mom_img, avg_mom_side, epoch, batch_no):
    #print("param_groups:", len(optimizer.param_groups))
    #print("state dict:", optimizer.state_dict()

    #momentum plot img.attention block
    sum_mom_img=0
    count_id_img = 0
    total_img_count = 0
    for name, param in model.named_parameters():
        if 'img.attention' in name:
            #print(id(param))
            sum_mom_img+=optimizer.state_dict()['state'][count_id_img]['exp_avg'].mean()
            total_img_count+=1
        count_id_img+=1
    avg_mom_block_img = (sum_mom_img/total_img_count).cpu().item()
    avg_mom_img.append(avg_mom_block_img)
    plot_momentum_flow(avg_mom_img, 'momentum_img.pdf', config_params)
    
    #momentum plot side.attention block
    sum_mom_side=0
    count_id_side = 0
    total_side_count = 0
    for name, param in model.named_parameters():
        if 'side.attention' in name:
            sum_mom_side+=optimizer.state_dict()['state'][count_id_side]['exp_avg'].mean()
            total_side_count+=1
        count_id_side+=1
    avg_mom_block_side = (sum_mom_side/total_side_count).cpu().item()
    avg_mom_side.append(avg_mom_block_side)
    plot_momentum_flow(avg_mom_side, 'momentum_side.pdf', config_params)
    
    wb_mom = op.load_workbook(os.path.join(config_params['path_to_output'], 'average_mom_epoch'+str(epoch)+'.xlsx'))
    sheet_mom = wb_mom['results']
    sheet_mom.append([batch_no, '+'.join(views_names), avg_mom_block_img, avg_mom_block_side])
    wb_mom.save((os.path.join(config_params['path_to_output'], 'average_mom_epoch'+str(epoch)+'.xlsx')))

    return case_mom_size, case_mom_dic_img, case_mom_dic_side, avg_mom_img, avg_mom_side